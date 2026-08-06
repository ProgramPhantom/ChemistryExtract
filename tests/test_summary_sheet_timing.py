from __future__ import annotations
import os
import openpyxl
from chemstractor.lib.reports.helpers import format_duration
from chemstractor.lib.reports.excel.create_excel import create_combined_excel

def test_format_duration():
    assert format_duration(None) == "N/A"
    assert format_duration("N/A") == "N/A"
    assert format_duration(0) == "0.00s"
    assert format_duration(15.25) == "15.25s"
    assert format_duration("15.25s") == "15.25s"
    assert format_duration(135.25) == "2m 15.25s"
    assert format_duration("135.25s") == "2m 15.25s"
    assert format_duration(3665.5) == "1h 1m 5s"
    assert format_duration(428558) == "4d 23h 2m 38s"

def test_summary_sheet_timing_rows(tmp_path):
    dest_file = os.path.join(tmp_path, "summary_timing_test.xlsx")
    
    mock_run_info = {
        "process_start_time": "2026-07-31 17:39:11",
        "process_end_time": "2026-07-31 17:41:26",
        "combine_start_time": "2026-08-05 16:41:00",
        "combine_end_time": "2026-08-05 16:41:50",
        "papers_duration_seconds": "135.00s",
        "combine_duration_seconds": "50.00s",
        "total_execution_seconds": 185.00,
        "duration_seconds": "185.00s"
    }

    combined_data = {
        "flory_entries": [],
        "mark_houwink_entries": [],
        "papers_summary": [],
        "run_info": mock_run_info
    }

    create_combined_excel(dest_file, combined_data)
    assert os.path.exists(dest_file)

    wb = openpyxl.load_workbook(dest_file)
    assert "Summary" in wb.sheetnames
    ws = wb["Summary"]

    # Read key-value pairs from Section 1 table
    rows_data = {}
    for r in range(6, 25):
        prop = ws.cell(row=r, column=1).value
        val = ws.cell(row=r, column=2).value
        if prop:
            rows_data[prop] = val

    assert rows_data.get("Process Stage Start Time") == "2026-07-31 17:39:11"
    assert rows_data.get("Process Stage End Time") == "2026-07-31 17:41:26"
    assert rows_data.get("Paper Processing Time") == "2m 15.00s"
    assert rows_data.get("Combine Stage Start Time") == "2026-08-05 16:41:00"
    assert rows_data.get("Combine Stage End Time") == "2026-08-05 16:41:50"
    assert rows_data.get("Combine Stage Execution Time") == "50.00s"
    assert rows_data.get("Total Execution Time") == "3m 5.00s"
