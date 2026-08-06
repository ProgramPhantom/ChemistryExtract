from __future__ import annotations
import os
import openpyxl
from chemstractor.lib.reports.excel.create_excel import create_combined_excel

def test_polymers_sheet_generation(tmp_path):
    dest_file = os.path.join(tmp_path, "combined_report_test.xlsx")
    
    mock_flory_entries = [
        {
            "polymer_name": "Polystyrene",
            "solvent_name": "Toluene",
            "temperature_k": 298.15,
            "c_value": -10.5,
            "v_value": 0.52,
            "failed_fields": "None",
            "source_paper": "Paper A",
            "table_name": "Table 1"
        },
        {
            "polymer_name": "Polystyrene",
            "solvent_name": "THF",
            "temperature_k": 298.15,
            "c_value": -10.8,
            "v_value": 0.55,
            "failed_fields": "None",
            "source_paper": "Paper B",
            "table_name": "Table 3"
        },
        {
            "polymer_name": "Poly(methyl methacrylate)",
            "solvent_name": "Acetone",
            "temperature_k": 293.15,
            "c_value": -9.9,
            "v_value": 0.48,
            "failed_fields": "None",
            "source_paper": "Paper C",
            "table_name": "Table 2"
        }
    ]
    
    combined_data = {
        "flory_entries": mock_flory_entries,
        "mark_houwink_entries": [],
        "papers_summary": []
    }
    
    create_combined_excel(dest_file, combined_data)
    assert os.path.exists(dest_file)
    
    wb = openpyxl.load_workbook(dest_file)
    assert "Results" in wb.sheetnames
    
    ws = wb["Results"]
    
    assert ws.freeze_panes == "A2"

    # Check Dropdown Label & Defaults (Right hand side: L2:M3)
    assert ws["L2"].value == "Select Polymer:"
    assert ws["M2"].value == "All"
    assert ws["L3"].value == "Select Solvent:"
    assert ws["M3"].value == "All"
    
    # Check Data Validation presence
    assert len(ws.data_validations.dataValidation) >= 2
    
    # Check Helper columns P (col 16) and Q (col 17)
    assert ws["P2"].value == "All"
    assert ws["Q2"].value == "All"
    poly_helpers = [ws.cell(row=r, column=16).value for r in range(2, 6)]
    solv_helpers = [ws.cell(row=r, column=17).value for r in range(2, 6)]
    
    assert "Polystyrene" in poly_helpers
    assert "Toluene" in solv_helpers

    # Check Dynamic Count Helper columns R and S
    assert "COUNTIFS" in str(ws["R2"].value)
    assert "COUNTIFS" in str(ws["S2"].value)

    # Check Table row 1 headers and row 2 data
    assert ws.cell(row=1, column=1).value == "Polymer (Clean)"
    assert ws.cell(row=1, column=8).value == "Source Paper"
    assert ws.cell(row=1, column=9).value == "Table"
    assert ws.cell(row=2, column=1).value == "Poly(methyl methacrylate)"
    assert ws.cell(row=2, column=8).value == "Paper C"
    assert ws.cell(row=2, column=9).value == "Table 2"
    
    # Check dynamic formula in column F & G
    formula_f = ws.cell(row=2, column=6).value
    assert formula_f.startswith("=IF(AND(OR(LEFT($M$2")
    assert "D2-E2*3, NA())" in formula_f

    # Native Excel chart present
    assert len(ws._charts) == 1
    chart = ws._charts[0]
    title_text = chart.title.tx.rich.p[0].r[0].t if (chart.title and chart.title.tx and chart.title.tx.rich) else str(chart.title)
    assert title_text == "Flory Calibration Curves (Interactive log-log Plot)"
