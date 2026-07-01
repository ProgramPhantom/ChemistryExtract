import os
import json
import csv
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import gc
import time
from chemstractor.lib.extractor import TableExtractor
from chemstractor.lib.categorisation import categorise_table
from chemstractor.lib.summariser import summarise_table_conditions, extract_paper_metadata
from chemstractor.models import pricing_matrix


class PDFProcessor:
    def __init__(self, pdf_path: str, output_dir: str = ".", model: str = "gemini-2.5-flash"):
        self.pdf_path = pdf_path
        self.model = model
        
        self.base_name = os.path.basename(pdf_path)
        self.base_no_ext = os.path.splitext(self.base_name)[0]
        
        # Output paths derived from output_dir
        self.output_dir = os.path.join(output_dir, self.base_no_ext)
        self.extract_dir = os.path.join(self.output_dir, "extract")
        self.tables_dir = os.path.join(self.extract_dir, "tables")
        
        self.clean_path = os.path.join(self.extract_dir, f"clean_{self.base_name}")
        self.parsed_md_path = os.path.join(self.extract_dir, "output.md")
        self.log_file_path = os.path.join(self.extract_dir, f"log_{self.base_name}.log")
        
        self.categorisation_dir = os.path.join(self.output_dir, "categorisation")
        self.summary_dir = os.path.join(self.output_dir, "summary")
        self.summary_json_path = os.path.join(self.summary_dir, "summary.json")
        
        # State
        self.extractor = None
        self.cat_data_list = []
        self.summarisation_data_list = []
        self.num_tables = 0
        self.cat_results = []
        self.sum_results = []
        self.metadata_res = None


    def _log_error(self, message: str):
        """Appends an error message to the log file."""
        try:
            with open(self.log_file_path, "a", encoding="utf-8") as lf:
                lf.write(f"PROCESSOR ERROR: {message}\n")
        except Exception:
            pass

    def extract(self):
        """Extracts content in-memory and yields status messages."""
        start_time = time.time()
        yield {"status": "working", "message": "Extracting text & tables..."}
        
        self.extractor = TableExtractor(self.pdf_path)
        self.num_tables = len(self.extractor.tables_markdown)

        elapsed_time = time.time() - start_time
        yield {
            "status": "complete",
            "message": "Extracted text & tables",
            "elapsed_time": elapsed_time,
            "num_tables": self.num_tables
        }

    
    def categorise(self):
        """Categorises each table in-memory and yields status events."""
        if not self.extractor:
            raise RuntimeError("Must call extract() before categorising tables.")
            
        start_time = time.time()
        yield {"status": "working", "message": "Categorising extracted tables..."}
        
        self.cat_results = []
        self.cat_data_list = []
        
        for i in range(self.num_tables):
            table_name = f"table{i + 1}.txt"
            
            yield {
                "status": "table_start",
                "table_idx": i,
                "table_name": table_name,
                "message": f"Categorising table {i + 1}/{self.num_tables}..."
            }
            
            table_text = self.extractor.tables_markdown[i]
            res = categorise_table(table_text, model=self.model)
            if res.success:
                status = "Contains" if res.contains_diffusion else "Does NOT contain"
                status_msg = f"{status} chemical diffusion coefficient data"
                self.cat_results.append((table_name, True, status_msg, res.usage_metadata))
                
                categorisation_data = {
                    "contains_scientific_data": res.contains_scientific_data,
                    "contains_diffusion_coeff": res.contains_diffusion_coeff,
                    "contains_polymer_diffusion_coeff": res.contains_polymer_diffusion_coeff,
                    "contains_diffusion": res.contains_diffusion
                }
                self.cat_data_list.append(categorisation_data)
                
                yield {
                    "status": "table_complete",
                    "table_idx": i,
                    "table_name": table_name,
                    "success": True,
                    "status_message": status_msg,
                    "usage_metadata": res.usage_metadata
                }
            else:
                status_msg = f"Failed: {res.error}"
                self.cat_results.append((table_name, False, status_msg, None))
                self.cat_data_list.append(None)
                yield {
                    "status": "table_complete",
                    "table_idx": i,
                    "table_name": table_name,
                    "success": False,
                    "status_message": status_msg,
                    "usage_metadata": None
                }
                
        elapsed_time = time.time() - start_time
        yield {
            "status": "complete",
            "message": "Categorised extracted tables",
            "elapsed_time": elapsed_time,
            "results": self.cat_results
        }

    def summarise(self):
        """Summarises each table in-memory and yields status events."""
        if not self.extractor:
            raise RuntimeError("Must call extract() before summarising tables.")
            
        start_time = time.time()
        yield {"status": "working", "message": "Extracting paper-level metadata..."}
        
        self.sum_results = []
        self.summarisation_data_list = []
        
        # Extract paper-level metadata
        metadata_res = extract_paper_metadata(self.extractor.parsed_markdown, model=self.model)
        self.metadata_res = metadata_res
        
        metadata_dict = {}
        metadata_error = None
        if metadata_res.success:
            metadata_dict = metadata_res.data.model_dump()
        else:
            metadata_error = metadata_res.error
            
        yield {
            "status": "metadata_complete",
            "success": metadata_res.success,
            "error": metadata_error,
            "usage_metadata": metadata_res.usage_metadata
        }
        
        for i in range(self.num_tables):
            table_name = f"table{i + 1}.txt"
            
            yield {
                "status": "table_start",
                "table_idx": i,
                "table_name": table_name,
                "message": f"Summarising table {i + 1}/{self.num_tables}..."
            }
            
            table_text = self.extractor.tables_markdown[i]
            res = summarise_table_conditions(table_text, model=self.model)
            
            if res.success:
                # Store experimental conditions only
                self.summarisation_data_list.append(res.data.model_dump())
                
                status_msg = "Successfully summarised"
                if metadata_error:
                    status_msg += f" (metadata extraction failed: {metadata_error})"
                self.sum_results.append((table_name, True, status_msg, res.usage_metadata))
                
                yield {
                    "status": "table_complete",
                    "table_idx": i,
                    "table_name": table_name,
                    "success": True,
                    "status_message": status_msg,
                    "usage_metadata": res.usage_metadata
                }
            else:
                status_msg = f"Failed: {res.error}"
                self.sum_results.append((table_name, False, status_msg, None))
                self.summarisation_data_list.append(None)
                yield {
                    "status": "table_complete",
                    "table_idx": i,
                    "table_name": table_name,
                    "success": False,
                    "status_message": status_msg,
                    "usage_metadata": None
                }
                
        elapsed_time = time.time() - start_time
        yield {
            "status": "complete",
            "message": "Summarised experimental conditions",
            "elapsed_time": elapsed_time,
            "results": self.sum_results
        }

    def create_excel(self) -> None:
        """Creates a beautifully formatted Excel document containing both JSON metadata/conditions and CSV table data."""
        wb = openpyxl.Workbook()
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])

        font_family = "Segoe UI"
        
        title_font = Font(name=font_family, size=16, bold=True, color="1F497D")
        title_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
        title_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        section_font = Font(name=font_family, size=11, bold=True, color="1F497D")
        section_fill = PatternFill(start_color="B8CCE4", end_color="B8CCE4", fill_type="solid")
        
        label_font = Font(name=font_family, size=10, bold=True, color="333333")
        val_font = Font(name=font_family, size=10, color="000000")
        desc_val_font = Font(name=font_family, size=10, italic=True, color="333333")
        
        table_header_font = Font(name=font_family, size=10, bold=True, color="FFFFFF")
        table_header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        table_header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        thin_border_side = Side(border_style="thin", color="D9D9D9")
        data_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
        thick_bottom = Border(bottom=Side(border_style="medium", color="1F497D"))
        
        left_align = Alignment(horizontal="left", vertical="center")
        right_align = Alignment(horizontal="right", vertical="center")
        wrap_left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
        # Load paper metadata from summary.json
        metadata = {}
        if os.path.exists(self.summary_json_path):
            try:
                with open(self.summary_json_path, 'r', encoding='utf-8') as mf:
                    metadata = json.load(mf)
            except Exception as e:
                self._log_error(f"Error loading metadata JSON {self.summary_json_path}: {e}")

        for i in range(self.num_tables):
            sheet_name = f"Table {i + 1}"
            ws = wb.create_sheet(title=sheet_name)
            ws.views.sheetView[0].showGridLines = True
            
            table_json_path = os.path.join(self.summary_dir, "tables", f"table{i + 1}.json")
            csv_path = os.path.join(self.tables_dir, "csv", f"table{i + 1}.csv")
            
            table_data = {}
            if os.path.exists(table_json_path):
                try:
                    with open(table_json_path, 'r', encoding='utf-8') as jf:
                        table_data = json.load(jf)
                except Exception as e:
                    self._log_error(f"Error loading table JSON {table_json_path}: {e}")
            
            title_text = metadata.get("title", self.base_no_ext)
            ws.merge_cells("A1:G2")
            title_cell = ws["A1"]
            title_cell.value = title_text
            title_cell.font = title_font
            title_cell.fill = title_fill
            title_cell.alignment = title_align
            
            for row in range(1, 3):
                for col in range(1, 8):
                    cell = ws.cell(row=row, column=col)
                    cell.fill = title_fill
                    
            ws.cell(row=4, column=1, value="PAPER METADATA & EXPERIMENTAL CONDITIONS").font = section_font
            ws.merge_cells("A4:G4")
            for col in range(1, 8):
                cell = ws.cell(row=4, column=col)
                cell.fill = section_fill
                cell.border = thick_bottom

            def write_metadata_row(r, label, value, is_description=False):
                ws.cell(row=r, column=1, value=label).font = label_font
                ws.cell(row=r, column=1).alignment = left_align
                ws.cell(row=r, column=2, value=value).font = desc_val_font if is_description else val_font
                ws.cell(row=r, column=2).alignment = wrap_left_align
                ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=7)
                for col in range(1, 8):
                    ws.cell(row=r, column=col).border = Border(bottom=thin_border_side)
            
            authors_val = ", ".join(metadata.get("authors", [])) if isinstance(metadata.get("authors"), list) else metadata.get("authors", "")
            write_metadata_row(5, "Authors", authors_val)
            write_metadata_row(6, "DOI", metadata.get("doi", ""))
            write_metadata_row(7, "Temperature", table_data.get("temperature", ""))
            write_metadata_row(8, "Pressure", table_data.get("pressure", ""))
            chemicals_val = ", ".join(table_data.get("chemicals", [])) if isinstance(table_data.get("chemicals"), list) else table_data.get("chemicals", "")
            write_metadata_row(9, "Chemicals", chemicals_val)
            
            desc_val = table_data.get("description", "")
            write_metadata_row(10, "Description", desc_val, is_description=True)
            ws.row_dimensions[10].height = 45
            
            other_stats = table_data.get("other_statistics", [])
            stats_str = ""
            if isinstance(other_stats, list):
                stats_str = ", ".join([f"{stat.get('name')}: {stat.get('value')}" for stat in other_stats if isinstance(stat, dict)])
            elif isinstance(other_stats, dict):
                stats_str = ", ".join([f"{k}: {v}" for k, v in other_stats.items()])
            write_metadata_row(11, "Other Stats", stats_str)

            cat_path = os.path.join(self.categorisation_dir, f"table{i + 1}.json")
            cat_data = {}
            if os.path.exists(cat_path):
                try:
                    with open(cat_path, 'r', encoding='utf-8') as cf:
                        cat_data = json.load(cf)
                except Exception as e:
                    print(f"Error loading categorisation JSON {cat_path}: {e}")

            curr_row = 13
            if cat_data:
                ws.cell(row=curr_row, column=1, value="TABLE CATEGORISATION").font = section_font
                ws.merge_cells(start_row=curr_row, start_column=1, end_row=curr_row, end_column=7)
                for col in range(1, 8):
                    cell = ws.cell(row=curr_row, column=col)
                    cell.fill = section_fill
                    cell.border = thick_bottom
                
                curr_row += 1
                def format_bool(val):
                    if val is True:
                        return "Yes"
                    if val is False:
                        return "No"
                    return ""
                write_metadata_row(curr_row, "Contains Scientific Data", format_bool(cat_data.get("contains_scientific_data")))
                curr_row += 1
                write_metadata_row(curr_row, "Contains Diffusion Coefficients", format_bool(cat_data.get("contains_diffusion_coeff")))
                curr_row += 1
                write_metadata_row(curr_row, "Contains Polymer Diffusion Coefficients", format_bool(cat_data.get("contains_polymer_diffusion_coeff")))
                
                curr_row += 2 # gap + dynamic next section start

            start_row = curr_row
            ws.cell(row=start_row, column=1, value="EXTRACTED TABLE DATA").font = section_font
            ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=7)
            for col in range(1, 8):
                cell = ws.cell(row=start_row, column=col)
                cell.fill = section_fill
                cell.border = thick_bottom
                
            csv_start_row = start_row + 1
            
            if os.path.exists(csv_path):
                try:
                    with open(csv_path, 'r', encoding='utf-8', newline='') as cf:
                        reader = csv.reader(cf)
                        csv_rows = list(reader)
                    
                    active_row_idx = 0
                    for csv_row in csv_rows:
                        # Skip empty rows (blank line or list containing only whitespace strings)
                        if not csv_row or all(val.strip() == '' for val in csv_row):
                            continue
                            
                        curr_row_idx = csv_start_row + active_row_idx
                        is_header = (active_row_idx == 0)
                        
                        for c_idx, val in enumerate(csv_row):
                            cell = ws.cell(row=curr_row_idx, column=c_idx + 1)
                            
                            parsed_val = val
                            try:
                                if "." in val:
                                    parsed_val = float(val)
                                else:
                                    parsed_val = int(val)
                            except ValueError:
                                pass
                                
                            cell.value = parsed_val
                            cell.border = data_border
                            
                            if is_header:
                                cell.font = table_header_font
                                cell.fill = table_header_fill
                                cell.alignment = table_header_align
                            else:
                                cell.font = val_font
                                if isinstance(parsed_val, (int, float)):
                                    cell.alignment = right_align
                                else:
                                    cell.alignment = left_align
                            active_row_idx += 1
                except Exception as e:
                    self._log_error(f"Error parsing CSV {csv_path}: {e}")
                    ws.cell(row=csv_start_row, column=1, value=f"Error loading CSV data: {e}").font = val_font
            else:
                ws.cell(row=csv_start_row, column=1, value="CSV data file not found.").font = val_font

            excluded_rows = {1, 2, 4, 10, start_row}
            if cat_data:
                excluded_rows.add(13)

            for col in ws.columns:
                max_len = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    if cell.row in excluded_rows:
                        continue
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
                ws.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 40)

        output_filename = f"{self.base_no_ext}_summary.xlsx"
        dest_path = os.path.join(self.summary_dir, output_filename)
        try:
            wb.save(dest_path)
        except Exception as e:
            self._log_error(f"Error saving Excel document {dest_path}: {e}")


    def save_logs(self):
        """Saves the extraction log file."""
        if not self.extractor:
            raise RuntimeError("Must call extract() before saving logs.")
        with open(self.log_file_path, "w", encoding="utf-8") as log_file:
            log_file.write(self.extractor.logs)

    def save_cleaned_pdf(self):
        """Saves the cleaned PDF byte content to the specified path."""
        if not self.extractor:
            raise RuntimeError("Must call extract() before saving cleaned PDF.")
        with open(self.clean_path, "wb") as clean_file:
            clean_file.write(self.extractor.clean_pdf_bytes)

    def save_summary_json(self):
        """Saves the paper metadata to the summary JSON file."""
        if self.metadata_res and self.metadata_res.success:
            os.makedirs(self.summary_dir, exist_ok=True)
            try:
                with open(self.summary_json_path, 'w', encoding='utf-8') as jf:
                    json.dump(self.metadata_res.data.model_dump(), jf, indent=2)
            except Exception as e:
                self._log_error(f"Error saving summary JSON {self.summary_json_path}: {e}")

    def save_outputs(self):
        """Saves the parsed markdown and extraction tables (text & csv)."""
        if not self.extractor:
            raise RuntimeError("Must call extract() before saving outputs.")
            
        # Save parsed markdown
        with open(self.parsed_md_path, "w", encoding="utf-8") as parsed_file:
            parsed_file.write(self.extractor.parsed_markdown)
            
        # Save text tables
        txt_dir = os.path.join(self.tables_dir, "txt")
        os.makedirs(txt_dir, exist_ok=True)
        for i, table_str in enumerate(self.extractor.tables_markdown):
            table_file_path = os.path.join(txt_dir, f"table{i + 1}.txt")
            with open(table_file_path, "w", encoding="utf-8") as table_file:
                table_file.write(table_str)
                
        # Save CSV tables
        csv_dir = os.path.join(self.tables_dir, "csv")
        os.makedirs(csv_dir, exist_ok=True)
        for i, csv_str in enumerate(self.extractor.tables_csv):
            csv_file_path = os.path.join(csv_dir, f"table{i + 1}.csv")
            with open(csv_file_path, "w", encoding="utf-8") as csv_file:
                csv_file.write(csv_str)
                
        # Save paper metadata summary if available
        self.save_summary_json()

    def save_all(self):
        """Saves all relevant content in-memory to the output folder structure."""
        if not self.extractor:
            raise RuntimeError("Must call extract() before saving.")
        
        # Create output directories
        os.makedirs(self.output_dir, exist_ok=True)
        os.makedirs(self.extract_dir, exist_ok=True)
        os.makedirs(self.tables_dir, exist_ok=True)
        
        # 1. Save extraction files
        self.save_cleaned_pdf()
        self.save_outputs()
        self.save_logs()
        
        # 2. Save categorisation JSONs if available
        if self.cat_data_list:
            os.makedirs(self.categorisation_dir, exist_ok=True)
            for i, cat_data in enumerate(self.cat_data_list):
                if cat_data is not None:
                    json_file_path = os.path.join(self.categorisation_dir, f"table{i + 1}.json")
                    with open(json_file_path, 'w', encoding='utf-8') as jf:
                        json.dump(cat_data, jf, indent=2)
                        
        # 3. Save summarisation JSONs if available
        if self.summarisation_data_list:
            tables_summary_dir = os.path.join(self.summary_dir, "tables")
            os.makedirs(tables_summary_dir, exist_ok=True)
            for i, sum_data in enumerate(self.summarisation_data_list):
                if sum_data is not None:
                    json_file_path = os.path.join(tables_summary_dir, f"table{i + 1}.json")
                    with open(json_file_path, 'w', encoding='utf-8') as jf:
                        json.dump(sum_data, jf, indent=2)
                        
        self.save_summary_json()
                        
        # 4. Save Excel summary if categorisation or summarisation was run
        if self.cat_data_list or self.summarisation_data_list:
            os.makedirs(self.summary_dir, exist_ok=True)
            self.create_excel()

    def cleanup(self):
        """Deload extractor instance and force garbage collection to release memory."""
        self.extractor = None
        gc.collect()
