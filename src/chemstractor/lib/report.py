import os
import sys
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, Reference
from openpyxl.drawing.line import LineProperties

def create_excel(
    dest_path: str,
    base_no_ext: str,
    metadata: dict,
    tables_data: list[dict],
    log_error_fn=None
) -> None:
    """Creates a beautifully formatted Excel document containing both JSON metadata/conditions and CSV table data."""
    def log_error(msg: str):
        if log_error_fn:
            log_error_fn(msg)
        else:
            print(f"REPORT ERROR: {msg}", file=sys.stderr)

    wb = openpyxl.Workbook()
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    font_family = "Segoe UI"
    
    title_font = Font(name=font_family, size=16, bold=True, color="1F497D")
    title_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
    title_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    section_font = Font(name=font_family, size=11, bold=True, color="1F497D")
    section_fill = PatternFill(start_color="B8CCE4", end_color="B8CCE4", fill_type="solid")
    
    label_font = Font(name=font_family, size=10, bold=True, color="333333")
    val_font = Font(name=font_family, size=10, color="000000")
    desc_val_font = Font(name=font_family, size=10, italic=True, color="333333")
    
    table_header_font = Font(name=font_family, size=10, bold=True, color="FFFFFF")
    table_header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    table_header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    thin_border_side = Side(border_style="thin", color="D9D9D9")
    data_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    thick_bottom = Border(bottom=Side(border_style="medium", color="1F497D"))
    
    left_align = Alignment(horizontal="left", vertical="center")
    right_align = Alignment(horizontal="right", vertical="center")
    wrap_left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)

    num_tables = len(tables_data)
    for i in range(num_tables):
        sheet_name = f"Table {i + 1}"
        ws = wb.create_sheet(title=sheet_name)
        ws.views.sheetView[0].showGridLines = True
        
        t_data = tables_data[i]
        table_data = t_data.get("table_data") or {}
        cat_data = t_data.get("cat_data") or {}
        csv_rows = t_data.get("csv_rows")
        csv_error = t_data.get("csv_error")
        
        title_text = metadata.get("title", base_no_ext)
        ws.merge_cells("A1:G2")
        title_cell = ws["A1"]
        title_cell.value = title_text
        title_cell.font = title_font
        title_cell.fill = title_fill
        title_cell.alignment = title_align
        
        for row in range(1, 3):
            for col in range(1, 8):
                cell = ws.cell(row=row, column=col)
                cell.fill = title_fill
                
        ws.cell(row=4, column=1, value="PAPER METADATA & EXPERIMENTAL CONDITIONS").font = section_font
        ws.merge_cells("A4:G4")
        for col in range(1, 8):
            cell = ws.cell(row=4, column=col)
            cell.fill = section_fill
            cell.border = thick_bottom

        def write_metadata_row(r, label, value, is_description=False):
            ws.cell(row=r, column=1, value=label).font = label_font
            ws.cell(row=r, column=1).alignment = left_align
            ws.cell(row=r, column=2, value=value).font = desc_val_font if is_description else val_font
            ws.cell(row=r, column=2).alignment = wrap_left_align
            ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=7)
            for col in range(1, 8):
                ws.cell(row=r, column=col).border = Border(bottom=thin_border_side)
        
        authors_val = ", ".join(metadata.get("authors", [])) if isinstance(metadata.get("authors"), list) else metadata.get("authors", "")
        write_metadata_row(5, "Authors", authors_val)
        write_metadata_row(6, "DOI", metadata.get("doi", ""))
        write_metadata_row(7, "Temperature", table_data.get("temperature", ""))
        write_metadata_row(8, "Pressure", table_data.get("pressure", ""))
        chemicals_val = ", ".join(table_data.get("chemicals", [])) if isinstance(table_data.get("chemicals"), list) else table_data.get("chemicals", "")
        write_metadata_row(9, "Chemicals", chemicals_val)
        
        desc_val = table_data.get("description", "")
        write_metadata_row(10, "Description", desc_val, is_description=True)
        ws.row_dimensions[10].height = 45
        
        other_stats = table_data.get("other_statistics", [])
        stats_str = ""
        if isinstance(other_stats, list):
            stats_str = ", ".join([f"{stat.get('name')}: {stat.get('value')}" for stat in other_stats if isinstance(stat, dict)])
        elif isinstance(other_stats, dict):
            stats_str = ", ".join([f"{k}: {v}" for k, v in other_stats.items()])
        write_metadata_row(11, "Other Stats", stats_str)

        section_rows = {1, 2, 4, 10}

        curr_row = 13
        if cat_data:
            ws.cell(row=curr_row, column=1, value="TABLE CATEGORISATION").font = section_font
            section_rows.add(curr_row)
            ws.merge_cells(start_row=curr_row, start_column=1, end_row=curr_row, end_column=7)
            for col in range(1, 8):
                cell = ws.cell(row=curr_row, column=col)
                cell.fill = section_fill
                cell.border = thick_bottom
            
            curr_row += 1
            def format_val(val):
                if val is True:
                    return "Yes"
                if val is False:
                    return "No"
                if isinstance(val, str):
                    return val
                return ""
            write_metadata_row(curr_row, "Contains Scientific Data", format_val(cat_data.get("contains_scientific_data")))
            curr_row += 1
            if "contains_diffusion_coeff" in cat_data:
                write_metadata_row(curr_row, "Contains Diffusion Coefficients", format_val(cat_data.get("contains_diffusion_coeff")))
                curr_row += 1
            else:
                write_metadata_row(curr_row, "Contains Raw Diffusion Data", format_val(cat_data.get("contains_raw_diffusion_data")))
                curr_row += 1
                write_metadata_row(curr_row, "Contains Mark-Houwink Parameters", format_val(cat_data.get("contains_mark_houwink_parameters")))
                curr_row += 1
                write_metadata_row(curr_row, "Contains Flory Parameters", format_val(cat_data.get("contains_flory_parameters")))
                curr_row += 1
            write_metadata_row(curr_row, "Contains Polymer Diffusion Coefficients", format_val(cat_data.get("contains_polymer_diffusion_coeff")))
            
            curr_row += 2 # gap + dynamic next section start

        # Interpretation Data (Flory Parameters & Plot)
        interpretation = t_data.get("interpretation") or {}
        flory_entries = interpretation.get("flory_entries", [])
        if flory_entries:
            ws.cell(row=curr_row, column=1, value="FLORY PARAMETERS (INTERPRETATION)").font = section_font
            section_rows.add(curr_row)
            ws.merge_cells(start_row=curr_row, start_column=1, end_row=curr_row, end_column=7)
            for col in range(1, 8):
                cell = ws.cell(row=curr_row, column=col)
                cell.fill = section_fill
                cell.border = thick_bottom
                
            curr_row += 1
            headers = ["Polymer", "Solvent", "c_value (log Constant)", "v_value (Scaling Exponent)"]
            for idx, h in enumerate(headers):
                cell = ws.cell(row=curr_row, column=idx + 1, value=h)
                cell.font = table_header_font
                cell.fill = table_header_fill
                cell.alignment = table_header_align
                cell.border = data_border
                
            start_table_row = curr_row + 1
            
            # Helper columns for plot data: Columns 5 to 7 (E to G)
            helper_header_font = Font(name=font_family, size=9, bold=True, color="7F7F7F")
            helper_val_font = Font(name=font_family, size=9, italic=True, color="7F7F7F")
            
            ws.cell(row=start_table_row - 1, column=5, value="Series").font = helper_header_font
            ws.cell(row=start_table_row - 1, column=5).alignment = left_align
            ws.cell(row=start_table_row - 1, column=5).border = data_border
            
            for col_idx, log_m_val in enumerate([3, 6]):
                c = 6 + col_idx
                cell = ws.cell(row=start_table_row - 1, column=c, value=log_m_val)
                cell.font = helper_header_font
                cell.alignment = right_align
                cell.border = data_border
                
            # Write entry rows
            for idx, entry in enumerate(flory_entries):
                r = start_table_row + idx
                
                # Main columns (A to D)
                ws.cell(row=r, column=1, value=entry.get("polymer_name")).font = val_font
                ws.cell(row=r, column=1).alignment = left_align
                ws.cell(row=r, column=1).border = data_border
                
                ws.cell(row=r, column=2, value=entry.get("solvent")).font = val_font
                ws.cell(row=r, column=2).alignment = left_align
                ws.cell(row=r, column=2).border = data_border
                
                ws.cell(row=r, column=3, value=entry.get("c_value")).font = val_font
                ws.cell(row=r, column=3).alignment = right_align
                ws.cell(row=r, column=3).border = data_border
                
                ws.cell(row=r, column=4, value=entry.get("v_value")).font = val_font
                ws.cell(row=r, column=4).alignment = right_align
                ws.cell(row=r, column=4).border = data_border
                
                # Series name (E)
                series_name = f"{entry.get('polymer_name')} in {entry.get('solvent')}"
                s_cell = ws.cell(row=r, column=5, value=series_name)
                s_cell.font = helper_val_font
                s_cell.alignment = left_align
                s_cell.border = data_border
                
                # Calculated points columns F and G
                for col_idx in range(2):
                    c = 6 + col_idx
                    log_m_val = 3 if col_idx == 0 else 6
                    # log(D) = c_value - v_value * log(M)
                    f_cell = ws.cell(row=r, column=c, value=f"=C{r}-D{r}*{log_m_val}")
                    f_cell.font = helper_val_font
                    f_cell.alignment = right_align
                    f_cell.border = data_border
                    
            end_table_row = start_table_row + len(flory_entries) - 1
            curr_row = end_table_row + 1
            
            # Create Line Chart
            try:
                chart = LineChart()
                chart.title = "Flory Calibration Curves (log-log Plot)"
                chart.style = 13
                chart.y_axis.title = "log(D / m² s⁻¹)"
                chart.x_axis.title = "log(M / g mol⁻¹)"
                
                # Data: columns E to G (5 to 7), rows start_table_row to end_table_row
                # (Notice min_row starts at start_table_row, NOT start_table_row - 1)
                data_ref = Reference(ws, min_col=5, max_col=7, min_row=start_table_row, max_row=end_table_row)
                # Categories: columns F to G (6 to 7), row start_table_row - 1
                cats_ref = Reference(ws, min_col=6, max_col=7, min_row=start_table_row - 1, max_row=start_table_row - 1)
                
                chart.add_data(data_ref, titles_from_data=True, from_rows=True)
                chart.set_categories(cats_ref)
                
                # Set a distinct color spectrum for each series
                colors = [
                    "1F497D",  # Dark Blue
                    "C0504D",  # Crimson
                    "9BBB59",  # Sage/Olive
                    "8064A2",  # Muted Purple
                    "F79646",  # Soft Orange
                    "4BACC6",  # Muted Teal
                    "E26B0A",  # Coral
                    "7030A0",  # Royal Violet
                    "00B0F0",  # Sky Blue
                    "B65708",  # Burnt Orange
                    "4F81BD",  # Medium Blue
                    "5F497A",  # Dark Purple
                ]
                for s_idx, series in enumerate(chart.series):
                    color = colors[s_idx % len(colors)]
                    series.graphicalProperties.line = LineProperties(solidFill=color)
                
                # Position the chart on the right side of the sheet
                ws.add_chart(chart, "K4")
            except Exception as e:
                log_error(f"Error creating openpyxl chart: {e}")
                
            curr_row += 2 # gap

        start_row = curr_row
        ws.cell(row=start_row, column=1, value="EXTRACTED TABLE DATA").font = section_font
        section_rows.add(start_row)
        ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=7)
        for col in range(1, 8):
            cell = ws.cell(row=start_row, column=col)
            cell.fill = section_fill
            cell.border = thick_bottom
            
        csv_start_row = start_row + 1
        
        if csv_rows is not None:
            active_row_idx = 0
            for csv_row in csv_rows:
                # Skip empty rows (blank line or list containing only whitespace strings)
                if not csv_row or all(val.strip() == '' for val in csv_row):
                    continue
                    
                curr_row_idx = csv_start_row + active_row_idx
                is_header = (active_row_idx == 0)
                
                for c_idx, val in enumerate(csv_row):
                    cell = ws.cell(row=curr_row_idx, column=c_idx + 1)
                    
                    parsed_val = val
                    try:
                        if "." in val:
                            parsed_val = float(val)
                        else:
                            parsed_val = int(val)
                    except ValueError:
                        pass
                        
                    cell.value = parsed_val
                    cell.border = data_border
                    
                    if is_header:
                        cell.font = table_header_font
                        cell.fill = table_header_fill
                        cell.alignment = table_header_align
                    else:
                        cell.font = val_font
                        if isinstance(parsed_val, (int, float)):
                            cell.alignment = right_align
                        else:
                            cell.alignment = left_align
                    active_row_idx += 1
        elif csv_error is not None:
            ws.cell(row=csv_start_row, column=1, value=f"Error loading CSV data: {csv_error}").font = val_font
        else:
            ws.cell(row=csv_start_row, column=1, value="CSV data file not found.").font = val_font

        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.row in section_rows:
                    continue
                if cell.value:
                    val_str = str(cell.value)
                    # Ignore formula cells starting with '=' from column width auto-fitting
                    if not val_str.startswith("="):
                        max_len = max(max_len, len(val_str))
            ws.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 40)

    try:
        wb.save(dest_path)
    except Exception as e:
        log_error(f"Error saving Excel document {dest_path}: {e}")


