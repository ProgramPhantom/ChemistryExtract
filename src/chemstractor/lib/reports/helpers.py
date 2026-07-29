from openpyxl.styles import PatternFill

# Error Severity & Fill Styling Configuration
ERROR_SEVERITY_MAP = {
    # Critical / High Severity (Red fill)
    "polymer_name": {"severity": "deselected", "color": "FFF2CC", "label": "Polymer Identification Failure"},
    "solvent_name": {"severity": "deselected", "color": "FFF2CC", "label": "Solvent Identification Failure"},
    
    # Warning / Medium Severity (Yellow fill)
    "c_value_missing": {"severity": "problem", "color": "FCE4D6", "label": "c_value Not Found"},
    "c_value_out_of_range": {"severity": "problem", "color": "FCE4D6", "label": "c_value Out of Range"},
    "v_value_missing": {"severity": "problem", "color": "FCE4D6", "label": "v_value Not Found"},
    "v_value_out_of_range": {"severity": "problem", "color": "FCE4D6", "label": "v_value Out of Range"},
    "K_value_missing": {"severity": "problem", "color": "FCE4D6", "label": "K_value Not Found"},
    "K_value_out_of_range": {"severity": "problem", "color": "FCE4D6", "label": "K_value Out of Range"},
    "a_value_missing": {"severity": "problem", "color": "FCE4D6", "label": "a_value Not Found"},
    "a_value_out_of_range": {"severity": "problem", "color": "FCE4D6", "label": "a_value Out of Range"},
    "D_out_of_range": {"severity": "problem", "color": "FCE4D6", "label": "D Value Out of Range"},
    "eta_out_of_range": {"severity": "problem", "color": "FCE4D6", "label": "log10([eta]) Out of Range"},
    
    # Low / Info Severity (Light Gray fill)
    "temperature_missing": {"severity": "warning", "color": "EDEDED", "label": "Temperature Missing/Invalid"}
}

SEVERITY_PRIORITY = {"deselected": 2, "problem": 3, "warning": 1}

def is_entry_failed(entry: dict) -> bool:
    """
    Returns True if the entry has any active 'deselected' or 'problem' severity failure flags.
    Entries with only 'warning' severity flags (e.g., temperature_missing) return False.
    """
    ff = entry.get("failed_fields")
    if isinstance(ff, dict):
        for field_key, is_failed in ff.items():
            if is_failed:
                meta = ERROR_SEVERITY_MAP.get(field_key, {})
                if meta.get("severity", "problem") in ("deselected", "problem"):
                    return True
        return False
    elif isinstance(ff, str):
        return ff != "None"
    return (
        entry.get("polymer_name") in (None, "", "N/A", "Invalid chemical") or
        entry.get("solvent_name") in (None, "", "N/A", "Invalid chemical") or
        entry.get("solvent") in (None, "", "N/A", "Invalid chemical")
    )

def get_entry_row_fill(failed_fields: dict) -> PatternFill | None:
    """Calculates Excel row fill color based on the highest severity active error in failed_fields."""
    if not isinstance(failed_fields, dict):
        return None
    
    highest_score = 0
    chosen_color = None

    for field_key, is_failed in failed_fields.items():
        if is_failed:
            meta = ERROR_SEVERITY_MAP.get(field_key, {"severity": "warning", "color": "FFF2CC"})
            score = SEVERITY_PRIORITY.get(meta.get("severity", "warning"), 1)
            if score > highest_score:
                highest_score = score
                chosen_color = meta.get("color", "FFF2CC")

    if chosen_color:
        return PatternFill(start_color=chosen_color, end_color=chosen_color, fill_type="solid")
    return None

def get_entry_errors_info(entry: dict) -> tuple[str, PatternFill | None]:
    """
    Constructs a single formatted error string containing all active error messages for the entry,
    and returns a PatternFill for the cell corresponding to the highest severity active error.
    """
    ff = entry.get("failed_fields")
    active_keys = []
    
    if isinstance(ff, dict):
        active_keys = [k for k, v in ff.items() if v]
    elif is_entry_failed(entry):
        active_keys = ["polymer_name"]

    if not active_keys:
        return "None", None

    labels = []
    highest_score = 0
    chosen_color = None

    for k in active_keys:
        meta = ERROR_SEVERITY_MAP.get(k, {"severity": "warning", "color": "FFF2CC", "label": k})
        labels.append(meta.get("label", k))
        score = SEVERITY_PRIORITY.get(meta.get("severity", "warning"), 1)
        if score > highest_score:
            highest_score = score
            chosen_color = meta.get("color", "FFF2CC")

    error_str = ", ".join(labels)
    fill = PatternFill(start_color=chosen_color, end_color=chosen_color, fill_type="solid") if chosen_color else None
    return error_str, fill

def get_entry_errors_and_warnings_info(entry: dict) -> tuple[str, PatternFill | None, str, PatternFill | None]:
    """
    Splits entry active messages into Errors (deselected/problem severity) and Warnings (warning severity).
    Returns (err_str, err_fill, warn_str, warn_fill).
    """
    ff = entry.get("failed_fields")
    active_keys = []
    
    if isinstance(ff, dict):
        active_keys = [k for k, v in ff.items() if v]
    elif is_entry_failed(entry):
        active_keys = ["polymer_name"]

    if not active_keys:
        return "None", None, "None", None

    err_labels = []
    warn_labels = []
    err_highest_score = 0
    err_chosen_color = None
    warn_highest_score = 0
    warn_chosen_color = None

    for k in active_keys:
        meta = ERROR_SEVERITY_MAP.get(k, {"severity": "problem", "color": "FFF2CC", "label": k})
        label = meta.get("label", k)
        sev = meta.get("severity", "problem")
        color = meta.get("color", "FFF2CC")
        score = SEVERITY_PRIORITY.get(sev, 1)

        if sev == "warning":
            warn_labels.append(label)
            if score > warn_highest_score:
                warn_highest_score = score
                warn_chosen_color = color
        else:
            err_labels.append(label)
            if score > err_highest_score:
                err_highest_score = score
                err_chosen_color = color

    err_str = ", ".join(err_labels) if err_labels else "None"
    err_fill = PatternFill(start_color=err_chosen_color, end_color=err_chosen_color, fill_type="solid") if err_chosen_color else None

    warn_str = ", ".join(warn_labels) if warn_labels else "None"
    warn_fill = PatternFill(start_color=warn_chosen_color, end_color=warn_chosen_color, fill_type="solid") if warn_chosen_color else None

    return err_str, err_fill, warn_str, warn_fill

def get_field_error(entry: dict, field_key: str) -> str:
    """Returns the failure reason for a specific field, or 'None' if no error."""
    ff = entry.get("failed_fields")
    if isinstance(ff, dict):
        if field_key == "polymer_name" and ff.get("polymer_name"):
            return "Invalid chemical"
        if field_key in ("solvent", "solvent_name") and ff.get("solvent_name"):
            return "Invalid chemical"
        if field_key == "temperature_k" and ff.get("temperature_missing"):
            return "Missing"
        if field_key == "c_value":
            if ff.get("c_value_missing"):
                return "Missing"
            if ff.get("c_value_out_of_range"):
                return "Out of range"
        if field_key == "v_value":
            if ff.get("v_value_missing"):
                return "Missing"
            if ff.get("v_value_out_of_range"):
                return "Out of range"
        if field_key == "K_value":
            if ff.get("K_value_missing"):
                return "Missing"
            if ff.get("K_value_out_of_range"):
                return "Out of range"
        if field_key == "a_value":
            if ff.get("a_value_missing"):
                return "Missing"
            if ff.get("a_value_out_of_range"):
                return "Out of range"
        if field_key == "general_err" and (ff.get("D_out_of_range") or ff.get("eta_out_of_range")):
            return "Out of range"
        return "None"
        
    field_errors = entry.get("field_errors")
    if isinstance(field_errors, dict):
        err = field_errors.get(field_key)
        if err:
            return err
    if field_key == "polymer_name" and entry.get("polymer_name") == "N/A":
        return "Invalid chemical"
    if field_key in ("solvent", "solvent_name") and (entry.get("solvent_name") == "N/A" or entry.get("solvent") == "N/A"):
        return "Invalid chemical"
    return "None"
