import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from chemstractor.lib.reports.helpers import is_entry_failed, get_entry_errors_and_warnings_info

def build_flory_sheet(wb, flory_entries: list, font_family: str = "Segoe UI") -> None:
    """Builds the Flory worksheet in the Excel workbook."""
    ws_flory = wb.create_sheet(title="Flory")
    ws_flory.views.sheetView[0].showGridLines = True
    
    table_header_font = Font(name=font_family, size=10, bold=True, color="FFFFFF")
    table_header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    table_header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    val_font = Font(name=font_family, size=10, color="000000")
    thin_border_side = Side(border_style="thin", color="D9D9D9")
    data_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    left_align = Alignment(horizontal="left", vertical="center")
    right_align = Alignment(horizontal="right", vertical="center")
    
    headers_flory = [
        "Source Paper", "Table", "Polymer (Original)", "Polymer (Clean)", 
        "Solvent (Original)", "Solvent (Clean)", "Temperature (K)", 
        "c Value (log Constant)", "c Transformation", "v Value (Scaling Exponent)", "v Transformation", 
        "Errors", "Warnings"
    ]
    
    for col_idx, h in enumerate(headers_flory):
        cell = ws_flory.cell(row=1, column=col_idx + 1, value=h)
        cell.font = table_header_font
        cell.fill = table_header_fill
        cell.alignment = table_header_align
        cell.border = data_border
        
    for val, col in [(3, 14), (6, 15)]:
        cell = ws_flory.cell(row=1, column=col, value=val)
        cell.font = table_header_font
        cell.fill = table_header_fill
        cell.alignment = table_header_align
        cell.border = data_border
        
    for row_idx, entry in enumerate(flory_entries):
        r = row_idx + 2
        ws_flory.cell(row=r, column=1, value=entry.get("source_paper", "")).font = val_font
        ws_flory.cell(row=r, column=2, value=entry.get("table_name", "")).font = val_font
        ws_flory.cell(row=r, column=3, value=entry.get("polymer_name_original", "")).font = val_font
        ws_flory.cell(row=r, column=4, value=entry.get("polymer_name", "")).font = val_font
        ws_flory.cell(row=r, column=5, value=entry.get("solvent_name_original", entry.get("solvent_original", ""))).font = val_font
        ws_flory.cell(row=r, column=6, value=entry.get("solvent_name", entry.get("solvent", ""))).font = val_font
        
        temp = entry.get("temperature_k")
        ws_flory.cell(row=r, column=7, value=float(temp) if isinstance(temp, (int, float)) else temp).font = val_font
        
        c_val = entry.get("c_value")
        ws_flory.cell(row=r, column=8, value=float(c_val) if isinstance(c_val, (int, float)) else c_val).font = val_font
        
        ws_flory.cell(row=r, column=9, value=entry.get("c_transformation", "")).font = val_font
        
        v_val = entry.get("v_value")
        ws_flory.cell(row=r, column=10, value=float(v_val) if isinstance(v_val, (int, float)) else v_val).font = val_font
        
        ws_flory.cell(row=r, column=11, value=entry.get("v_transformation", "")).font = val_font
        
        err_str, err_fill, warn_str, warn_fill = get_entry_errors_and_warnings_info(entry)
        err_cell = ws_flory.cell(row=r, column=12, value=err_str)
        err_cell.font = val_font
        if err_fill:
            err_cell.fill = err_fill
            
        warn_cell = ws_flory.cell(row=r, column=13, value=warn_str)
        warn_cell.font = val_font
        if warn_fill:
            warn_cell.fill = warn_fill
        
        if isinstance(c_val, (int, float)) and isinstance(v_val, (int, float)):
            ws_flory.cell(row=r, column=14, value=f"=H{r}-J{r}*3").font = val_font
            ws_flory.cell(row=r, column=15, value=f"=H{r}-J{r}*6").font = val_font
            
        for c in [1, 2, 3, 4, 5, 6, 9, 11, 12, 13]:
            cell = ws_flory.cell(row=r, column=c)
            cell.alignment = left_align
            cell.border = data_border
        for c in [7, 8, 10, 14, 15]:
            cell = ws_flory.cell(row=r, column=c)
            cell.alignment = right_align
            cell.border = data_border
            
    for col in ws_flory.columns:
        vals = [str(cell.value or '') for cell in col]
        max_len = max(len(v) for v in vals) if vals else 10
        col_letter = get_column_letter(col[0].column)
        ws_flory.column_dimensions[col_letter].width = max(max_len + 3, 12)
