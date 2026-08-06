from datetime import datetime
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from chemstractor.lib.reports.helpers import is_entry_failed, ERROR_SEVERITY_MAP, format_duration

def build_summary_sheet(
    wb,
    combined_data: dict,
    failure_reasons_count: dict,
    severity_counts: dict,
    total_collected: int,
    font_family: str = "Segoe UI"
) -> None:
    """Builds the Summary worksheet in the Excel workbook."""
    ws_sum = wb.create_sheet(title="Summary")
    ws_sum.views.sheetView[0].showGridLines = True

    # Styles
    table_header_font = Font(name=font_family, size=10, bold=True, color="FFFFFF")
    table_header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    table_header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    section_title_font = Font(name=font_family, size=11, bold=True, color="1F497D")
    val_font = Font(name=font_family, size=10, color="000000")
    thin_border_side = Side(border_style="thin", color="D9D9D9")
    data_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    left_align = Alignment(horizontal="left", vertical="center")
    right_align = Alignment(horizontal="right", vertical="center")

    flory_entries = combined_data.get("flory_entries", [])
    mh_entries = combined_data.get("mark_houwink_entries", [])
    papers_summary = combined_data.get("papers_summary", [])
    run_info = combined_data.get("run_info", {})

    # Main Header Banner
    ws_sum.merge_cells("A1:B1")
    title_cell = ws_sum.cell(row=1, column=1, value="CHEMSTRACTOR BATCH PROCESSING REPORT")
    title_cell.font = Font(name=font_family, size=13, bold=True, color="FFFFFF")
    title_cell.fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws_sum.row_dimensions[1].height = 28
    
    ws_sum.merge_cells("A2:B2")
    sub_cell = ws_sum.cell(row=2, column=1, value=f"Report Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    sub_cell.font = Font(name=font_family, size=9.5, italic=True, color="595959")
    sub_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws_sum.row_dimensions[2].height = 18

    current_row = 4

    # --- Section 1: Run & Environment Information ---
    ws_sum.cell(row=current_row, column=1, value="1. Run Execution & System Information").font = section_title_font
    current_row += 1

    headers_meta = ["Property", "Details"]
    for c_idx, h in enumerate(headers_meta):
        cell = ws_sum.cell(row=current_row, column=c_idx + 1, value=h)
        cell.font = table_header_font
        cell.fill = table_header_fill
        cell.alignment = table_header_align
        cell.border = data_border
    current_row += 1

    total_tables_checked = sum(p.get("total_tables", 0) for p in papers_summary)
    total_tables_selected = sum(p.get("selected_tables", 0) for p in papers_summary)

    metadata_rows = [
        ("Process Stage Start Time", run_info.get("process_start_time", run_info.get("start_time", "N/A"))),
        ("Process Stage End Time", run_info.get("process_end_time", "N/A")),
        ("Paper Processing Time", format_duration(run_info.get("papers_duration_seconds"))),
        ("Combine Stage Start Time", run_info.get("combine_start_time", "N/A")),
        ("Combine Stage End Time", run_info.get("combine_end_time", run_info.get("end_time", "N/A"))),
        ("Combine Stage Execution Time", format_duration(run_info.get("combine_duration_seconds"))),
        ("Total Execution Time", format_duration(run_info.get("total_execution_seconds", run_info.get("duration_seconds")))),
        ("AI Model Selected", run_info.get("model_used", "N/A")),
        ("Input Directory / Run", run_info.get("input_directory", "N/A")),
        ("Total Papers Inputted", run_info.get("total_papers_inputted", run_info.get("num_papers", len(papers_summary)))),
        ("Papers Selected", run_info.get("num_papers", len(papers_summary))),
        ("Total Tables Inspected", run_info.get("total_tables", total_tables_checked)),
        ("Tables Selected for Extraction", run_info.get("selected_tables", total_tables_selected)),
        ("Operating System", run_info.get("os", "N/A")),
        ("CPU Architecture & Cores", f"{run_info.get('architecture', '')} ({run_info.get('cpu_cores', '')} Cores)".strip() if run_info.get("architecture") else "N/A"),
        ("Python Environment", run_info.get("python_version", "N/A")),
        ("Host Machine Name", run_info.get("node_hostname", "N/A"))
    ]

    for item_label, item_val in metadata_rows:
        c1 = ws_sum.cell(row=current_row, column=1, value=item_label)
        c1.font = val_font
        c1.alignment = left_align
        c1.border = data_border

        c2 = ws_sum.cell(row=current_row, column=2, value=str(item_val))
        c2.font = val_font
        c2.alignment = left_align if isinstance(item_val, str) else right_align
        c2.border = data_border
        current_row += 1

    current_row += 1  # Spacing

    # --- Section 2: Extraction Statistics & KPIs ---
    ws_sum.cell(row=current_row, column=1, value="2. Extraction & Homogenisation Statistics").font = section_title_font
    current_row += 1

    headers_sum_kpi = ["Metric", "Value"]
    for c_idx, h in enumerate(headers_sum_kpi):
        cell = ws_sum.cell(row=current_row, column=c_idx + 1, value=h)
        cell.font = table_header_font
        cell.fill = table_header_fill
        cell.alignment = table_header_align
        cell.border = data_border
    current_row += 1

    all_entries = flory_entries + mh_entries
    selected_data_points = 0
    deselected_data_points = 0
    for entry in all_entries:
        if is_entry_failed(entry):
            deselected_data_points += 1
        else:
            selected_data_points += 1

    success_rate = (selected_data_points / total_collected * 100) if total_collected > 0 else 0.0

    kpis = [
        ("Total Data Points Collected", total_collected),
        ("Flory Data Points", len(flory_entries)),
        ("Mark-Houwink Data Points", len(mh_entries)),
        ("Selected Data Points (Successful)", selected_data_points),
        ("Deselected Data Points (Rejected)", deselected_data_points),
        ("Issue Level: Deselected (Critical Failures)", severity_counts.get("deselected", 0)),
        ("Issue Level: Problem (Parameter Errors)", severity_counts.get("problem", 0)),
        ("Issue Level: Warning (Minor Warnings)", severity_counts.get("warning", 0)),
    ]

    for m, v in kpis:
        c1 = ws_sum.cell(row=current_row, column=1, value=m)
        c1.font = val_font
        c1.alignment = left_align
        c1.border = data_border
        
        c2 = ws_sum.cell(row=current_row, column=2, value=v)
        c2.font = val_font
        c2.alignment = right_align
        c2.border = data_border
        current_row += 1

    current_row += 1  # Spacing

    # --- Section 3: Failure Reasons Breakdown ---
    ws_sum.cell(row=current_row, column=1, value="3. Failure Reasons Breakdown").font = section_title_font
    current_row += 1

    headers_sum_fail = ["Failure Reason / Field", "Count"]
    for c_idx, h in enumerate(headers_sum_fail):
        cell = ws_sum.cell(row=current_row, column=c_idx + 1, value=h)
        cell.font = table_header_font
        cell.fill = table_header_fill
        cell.alignment = table_header_align
        cell.border = data_border
    current_row += 1

    if failure_reasons_count:
        sorted_reasons = sorted(failure_reasons_count.items(), key=lambda x: x[1], reverse=True)
        for reason, count in sorted_reasons:
            c1 = ws_sum.cell(row=current_row, column=1, value=reason)
            c1.font = val_font
            c1.alignment = left_align
            c1.border = data_border

            c2 = ws_sum.cell(row=current_row, column=2, value=count)
            c2.font = val_font
            c2.alignment = right_align
            c2.border = data_border
            current_row += 1
    else:
        c1 = ws_sum.cell(row=current_row, column=1, value="None (All extractions succeeded)")
        c1.font = val_font
        c1.alignment = left_align
        c1.border = data_border
        c2 = ws_sum.cell(row=current_row, column=2, value=0)
        c2.font = val_font
        c2.alignment = right_align
        c2.border = data_border
        current_row += 1

    for col in ws_sum.columns:
        vals = [str(cell.value or '') for cell in col]
        max_len = max(len(v) for v in vals) if vals else 10
        col_letter = get_column_letter(col[0].column)
        ws_sum.column_dimensions[col_letter].width = max(max_len + 4, 18)
