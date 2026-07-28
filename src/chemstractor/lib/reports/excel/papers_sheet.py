from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

def build_papers_sheet(wb, papers_summary: list, font_family: str = "Segoe UI") -> None:
    """Builds the Papers worksheet in the Excel workbook."""
    ws_papers = wb.create_sheet(title="Papers")
    ws_papers.views.sheetView[0].showGridLines = True

    table_header_font = Font(name=font_family, size=10, bold=True, color="FFFFFF")
    table_header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    table_header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    val_font = Font(name=font_family, size=10, color="000000")
    thin_border_side = Side(border_style="thin", color="D9D9D9")
    data_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    left_align = Alignment(horizontal="left", vertical="center")
    right_align = Alignment(horizontal="right", vertical="center")

    headers_papers = [
        "Source Paper", "Total Tables", "Selected Tables", 
        "Flory Data Points", "Mark-Houwink Data Points", "Failed Data Points"
    ]
    for col_idx, h in enumerate(headers_papers):
        cell = ws_papers.cell(row=1, column=col_idx + 1, value=h)
        cell.font = table_header_font
        cell.fill = table_header_fill
        cell.alignment = table_header_align
        cell.border = data_border

    for row_idx, p_info in enumerate(papers_summary):
        r = row_idx + 2
        ws_papers.cell(row=r, column=1, value=p_info.get("source_paper", "")).font = val_font
        ws_papers.cell(row=r, column=2, value=p_info.get("total_tables", 0)).font = val_font
        ws_papers.cell(row=r, column=3, value=p_info.get("selected_tables", 0)).font = val_font
        ws_papers.cell(row=r, column=4, value=p_info.get("flory_count", 0)).font = val_font
        ws_papers.cell(row=r, column=5, value=p_info.get("mh_count", 0)).font = val_font
        ws_papers.cell(row=r, column=6, value=p_info.get("failed_count", 0)).font = val_font

        ws_papers.cell(row=r, column=1).alignment = left_align
        ws_papers.cell(row=r, column=1).border = data_border
        for c in range(2, 7):
            ws_papers.cell(row=r, column=c).alignment = right_align
            ws_papers.cell(row=r, column=c).border = data_border

    for col in ws_papers.columns:
        vals = [str(cell.value or '') for cell in col]
        max_len = max(len(v) for v in vals) if vals else 10
        col_letter = get_column_letter(col[0].column)
        ws_papers.column_dimensions[col_letter].width = max(max_len + 3, 14)
