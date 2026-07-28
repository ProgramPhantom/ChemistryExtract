from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from chemstractor.lib.reports.helpers import ERROR_SEVERITY_MAP, SEVERITY_PRIORITY, is_entry_failed

def build_failures_sheet(wb, flory_entries: list, mh_entries: list, font_family: str = "Segoe UI") -> None:
    """Builds the Failures worksheet in the Excel workbook."""
    ws_fail = wb.create_sheet(title="Failures")
    ws_fail.views.sheetView[0].showGridLines = True
    
    table_header_font = Font(name=font_family, size=10, bold=True, color="FFFFFF")
    table_header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    table_header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    val_font = Font(name=font_family, size=10, color="000000")
    thin_border_side = Side(border_style="thin", color="D9D9D9")
    data_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    left_align = Alignment(horizontal="left", vertical="center")
    
    headers_fail = ["Source Paper", "Table", "Type", "Active Errors", "Primary Blame / Severity"]
    for col_idx, h in enumerate(headers_fail):
        cell = ws_fail.cell(row=1, column=col_idx + 1, value=h)
        cell.font = table_header_font
        cell.fill = table_header_fill
        cell.alignment = table_header_align
        cell.border = data_border
        
    failed_all_entries = []
    for entry in flory_entries:
        if is_entry_failed(entry):
            failed_all_entries.append(("Flory", entry))
    for entry in mh_entries:
        if is_entry_failed(entry):
            failed_all_entries.append(("Mark-Houwink", entry))

    for row_idx, (entry_type, entry) in enumerate(failed_all_entries):
        r = row_idx + 2
        ff = entry.get("failed_fields", {})
        active_errs = [k for k, v in ff.items() if v] if isinstance(ff, dict) else ["failed"]
        
        highest_score = 0
        primary_blame = "Error"
        fill_color = "FFF2CC"
        if isinstance(ff, dict):
            for k in active_errs:
                meta = ERROR_SEVERITY_MAP.get(k, {})
                score = SEVERITY_PRIORITY.get(meta.get("severity", "warning"), 1)
                if score > highest_score:
                    highest_score = score
                    primary_blame = meta.get("label", k)
                    fill_color = meta.get("color", "FFF2CC")

        ws_fail.cell(row=r, column=1, value=entry.get("source_paper", "")).font = val_font
        ws_fail.cell(row=r, column=2, value=entry.get("table_name", "")).font = val_font
        ws_fail.cell(row=r, column=3, value=entry_type).font = val_font
        ws_fail.cell(row=r, column=4, value=", ".join(active_errs)).font = val_font
        ws_fail.cell(row=r, column=5, value=primary_blame).font = val_font
        
        row_fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
        for c in range(1, 6):
            cell = ws_fail.cell(row=r, column=c)
            cell.alignment = left_align
            cell.border = data_border
            cell.fill = row_fill
            
    for col in ws_fail.columns:
        vals = [str(cell.value or '') for cell in col]
        max_len = max(len(v) for v in vals) if vals else 10
        col_letter = get_column_letter(col[0].column)
        ws_fail.column_dimensions[col_letter].width = max(max_len + 3, 12)
