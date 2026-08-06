from __future__ import annotations
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.chart import LineChart, Reference, Series
from chemstractor.lib.reports.helpers import is_entry_failed

def try_parse_float(val) -> float | None:
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)
    if isinstance(val, str):
        try:
            return float(val.strip())
        except ValueError:
            return None
    return None

def is_flory_entry_valid(entry: dict) -> bool:
    """Checks if a Flory entry is valid for inclusion in the Polymers sheet."""
    c_val = try_parse_float(entry.get("c_value"))
    v_val = try_parse_float(entry.get("v_value"))
    p_name = entry.get("polymer_name")
    s_name = entry.get("solvent_name", entry.get("solvent"))

    if c_val is None or v_val is None:
        return False
    if p_name in (None, "", "N/A", "Invalid chemical") or s_name in (None, "", "N/A", "Invalid chemical"):
        return False

    ff = entry.get("failed_fields")
    if isinstance(ff, dict):
        # Temperature missing is a warning and should be included
        critical_failures = [k for k, v in ff.items() if v and k != "temperature_missing"]
        if critical_failures:
            return False
    elif isinstance(ff, str) and ff != "None":
        return False

    return True

def build_polymers_sheet(wb, flory_entries: list, font_family: str = "Segoe UI") -> None:
    """Builds the interactive Polymers worksheet in the Excel workbook."""
    ws_poly = wb.create_sheet(title="Results")
    ws_poly.views.sheetView[0].showGridLines = True

    # Styling definitions
    title_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
    title_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    table_header_font = Font(name=font_family, size=10, bold=True, color="FFFFFF")
    table_header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    table_header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    label_font = Font(name=font_family, size=10, bold=True, color="1F497D")
    val_font = Font(name=font_family, size=10, color="000000")
    
    thin_border_side = Side(border_style="thin", color="D9D9D9")
    data_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    
    left_align = Alignment(horizontal="left", vertical="center")
    right_align = Alignment(horizontal="right", vertical="center")
    center_align = Alignment(horizontal="center", vertical="center")

    # Filter out failed entries
    valid_flory_poly = [entry for entry in flory_entries if is_flory_entry_valid(entry)]

    # Sort Flory entries polymer-centric
    sorted_flory_poly = sorted(
        valid_flory_poly, 
        key=lambda x: (
            (x.get("polymer_name") or "").lower(),
            (x.get("solvent_name") or x.get("solvent") or "").lower()
        )
    )

    # Gather unique polymers and solvents for dropdown lists
    unique_polymers = ["All"] + sorted(list({e.get("polymer_name") for e in sorted_flory_poly if e.get("polymer_name")}))
    unique_solvents = ["All"] + sorted(list({e.get("solvent_name", e.get("solvent")) for e in sorted_flory_poly if e.get("solvent_name", e.get("solvent"))}))

    # Row positioning definitions (Data table moved 50 lines lower: header at 69, data at 70)
    header_row = 69
    start_data_row = 70
    end_data_row = start_data_row + len(sorted_flory_poly) - 1 if sorted_flory_poly else start_data_row

    data_range_a = f"$A$70:$A${max(end_data_row, 70)}"
    data_range_b = f"$B$70:$B${max(end_data_row, 70)}"

    sel_poly_clean = "IF(OR(LEFT($C$3, 3)=\"All\", $C$3=\"\"), \"*\", LEFT($C$3, IF(ISNUMBER(SEARCH(\" (\", $C$3)), SEARCH(\" (\", $C$3)-1, LEN($C$3))))"
    sel_solv_clean = "IF(OR(LEFT($C$4, 3)=\"All\", $C$4=\"\"), \"*\", LEFT($C$4, IF(ISNUMBER(SEARCH(\" (\", $C$4)), SEARCH(\" (\", $C$4)-1, LEN($C$4))))"

    # Helper Column L: Raw Unique Polymers
    for idx, p in enumerate(unique_polymers, start=2):
        ws_poly.cell(row=idx, column=12, value=p).font = val_font

    # Helper Column M: Raw Unique Solvents
    for idx, s in enumerate(unique_solvents, start=2):
        ws_poly.cell(row=idx, column=13, value=s).font = val_font

    # Helper Column N: Dynamic Solvent List with Counts (Evaluating selected Polymer C3)
    ws_poly.cell(row=2, column=14, value=f'=M2 & " (" & COUNTIFS({data_range_a}, {sel_poly_clean}) & ")"').font = val_font
    for idx, s in enumerate(unique_solvents[1:], start=3):
        ws_poly.cell(row=idx, column=14, value=f'=M{idx} & " (" & COUNTIFS({data_range_a}, {sel_poly_clean}, {data_range_b}, M{idx}) & ")"').font = val_font

    # Helper Column O: Dynamic Polymer List with Counts (Evaluating selected Solvent C4)
    ws_poly.cell(row=2, column=15, value=f'=L2 & " (" & COUNTIFS({data_range_b}, {sel_solv_clean}) & ")"').font = val_font
    for idx, p in enumerate(unique_polymers[1:], start=3):
        ws_poly.cell(row=idx, column=15, value=f'=L{idx} & " (" & COUNTIFS({data_range_a}, L{idx}, {data_range_b}, {sel_solv_clean}) & ")"').font = val_font

    # Hide helper columns J, L, M, N, O
    for c_letter in ("J", "L", "M", "N", "O"):
        ws_poly.column_dimensions[c_letter].hidden = True

    # 1. Interactive Control Panel (Rows 2 - 5, Columns B & C)
    ws_poly.merge_cells("B2:C2")
    ctrl_title_cell = ws_poly.cell(row=2, column=2, value="Data explorer")
    ctrl_title_cell.font = title_font
    ctrl_title_cell.fill = title_fill
    ctrl_title_cell.alignment = center_align
    ctrl_title_cell.border = data_border
    ws_poly.cell(row=2, column=3).border = data_border

    # Select Polymer row
    lbl_poly = ws_poly.cell(row=3, column=2, value="Select Polymer:")
    lbl_poly.font = label_font
    lbl_poly.alignment = Alignment(horizontal="right", vertical="center")
    lbl_poly.border = data_border

    val_poly = ws_poly.cell(row=3, column=3, value="All")
    val_poly.font = val_font
    val_poly.alignment = left_align
    val_poly.border = data_border

    # Add Polymer Data Validation (Pointing to Column O)
    dv_poly = DataValidation(
        type="list", 
        formula1=f"=Polymers!$O$2:$O${len(unique_polymers)+1}", 
        allow_blank=True
    )
    ws_poly.add_data_validation(dv_poly)
    dv_poly.add(val_poly)

    # Select Solvent row
    lbl_solv = ws_poly.cell(row=4, column=2, value="Select Solvent:")
    lbl_solv.font = label_font
    lbl_solv.alignment = Alignment(horizontal="right", vertical="center")
    lbl_solv.border = data_border

    val_solv = ws_poly.cell(row=4, column=3, value="All")
    val_solv.font = val_font
    val_solv.alignment = left_align
    val_solv.border = data_border

    # Add Solvent Data Validation (Pointing to Column N)
    dv_solv = DataValidation(
        type="list", 
        formula1=f"=Polymers!$N$2:$N${len(unique_solvents)+1}", 
        allow_blank=True
    )
    ws_poly.add_data_validation(dv_solv)
    dv_solv.add(val_solv)

    # Ensure top row is not frozen/sticky
    ws_poly.freeze_panes = None


    # Python in Excel Matplotlib plot formula placed in cell E2
    py_plot_code = (
        '=PY("'
        'import matplotlib.pyplot as plt\n'
        'poly = xl(\'C3\')\n'
        'solv = xl(\'C4\')\n'
        'poly_clean = poly.split(\' (\')[0] if poly and str(poly) != \'All\' and not str(poly).startswith(\'All\') else None\n'
        'solv_clean = solv.split(\' (\')[0] if solv and str(solv) != \'All\' and not str(solv).startswith(\'All\') else None\n'
        f'df = xl(\'A{header_row}:I{max(end_data_row, start_data_row)}\', headers=True)\n'
        'fig, ax = plt.subplots(figsize=(7.5, 5.0), dpi=300)\n'
        'colors = [\'#1f77b4\', \'#ff7f0e\', \'#2ca02c\', \'#d62728\', \'#9467bd\', \'#8c564b\', \'#e377c2\', \'#7f7f7f\', \'#bcbd22\', \'#17becf\', \'#319795\', \'#d69e2e\', \'#805ad5\', \'#dd6b20\', \'#3182ce\']\n'
        'valid_count = 0\n'
        'for idx, row in df.iterrows():\n'
        '    p = str(row[\'Polymer (Clean)\'])\n'
        '    s = str(row[\'Solvent (Clean)\'])\n'
        '    c_val = row[\'c Value (log Constant)\']\n'
        '    v_val = row[\'v Value (Scaling Exponent)\']\n'
        '    if poly_clean and p != poly_clean:\n'
        '        continue\n'
        '    if solv_clean and s != solv_clean:\n'
        '        continue\n'
        '    try:\n'
        '        c = float(c_val)\n'
        '        v = float(v_val)\n'
        '        y3 = c - v * 3\n'
        '        y6 = c - v * 6\n'
        '        if poly_clean and not solv_clean:\n'
        '            label = s\n'
        '        elif solv_clean and not poly_clean:\n'
        '            label = p\n'
        '        else:\n'
        '            label = f\'{p} in {s}\'\n'
        '        color = colors[valid_count % len(colors)]\n'
        '        ax.plot([3, 6], [y3, y6], label=label, color=color, linewidth=2.0, alpha=0.9)\n'
        '        valid_count += 1\n'
        '    except (ValueError, TypeError):\n'
        '        continue\n'
        'if valid_count > 0:\n'
        '    ax.set_title(\'Flory Calibration Curves (Interactive log-log Plot)\', pad=12)\n'
        '    ax.set_xlabel(\'log10(M / g mol-1)\')\n'
        '    ax.set_ylabel(\'log10(D / m2 s-1)\')\n'
        '    ax.set_xticks([3, 4, 5, 6])\n'
        '    ax.grid(True, linestyle=\'--\', alpha=0.6, color=\'#cbd5e0\')\n'
        '    ax.spines[\'top\'].set_visible(False)\n'
        '    ax.spines[\'right\'].set_visible(False)\n'
        '    ax.legend(bbox_to_anchor=(1.04, 1), loc=\'upper left\', frameon=True, facecolor=\'#f7fafc\', edgecolor=\'#e2e8f0\', fontsize=8.5)\n'
        'plt.gcf()")'
    )
    ws_poly.cell(row=2, column=5, value=py_plot_code)

    # 2. Main Data Table Header (Row 69)
    headers_poly = [
        "Polymer (Clean)", "Solvent (Clean)", "Temperature (K)", 
        "c Value (log Constant)", "v Value (Scaling Exponent)", "3", "6",
        "Source Paper", "Table"
    ]
    for col_idx, h in enumerate(headers_poly, start=1):
        cell = ws_poly.cell(row=header_row, column=col_idx, value=h if not h.isdigit() else int(h))
        cell.font = table_header_font
        cell.fill = table_header_fill
        cell.alignment = table_header_align
        cell.border = data_border

    # Fill Main Data Table rows
    y_vals = []
    for row_offset, entry in enumerate(sorted_flory_poly):
        r = start_data_row + row_offset
        ws_poly.cell(row=r, column=1, value=entry.get("polymer_name", "")).font = val_font
        ws_poly.cell(row=r, column=2, value=entry.get("solvent_name", entry.get("solvent", ""))).font = val_font

        temp = entry.get("temperature_k")
        ws_poly.cell(row=r, column=3, value=float(temp) if isinstance(temp, (int, float)) else temp).font = val_font

        c_val = entry.get("c_value")
        ws_poly.cell(row=r, column=4, value=float(c_val) if isinstance(c_val, (int, float)) else c_val).font = val_font

        v_val = entry.get("v_value")
        ws_poly.cell(row=r, column=5, value=float(v_val) if isinstance(v_val, (int, float)) else v_val).font = val_font

        # Dynamic formula: Evaluate to numeric log value if matching Polymer and Solvent selection, else NA()
        formula_filter = (
            f"AND("
            f"OR(LEFT($C$3, 3)=\"All\", $C$3=\"\", $C$3=A{r}, AND(LEFT($C$3, LEN(A{r}))=A{r}, MID($C$3, LEN(A{r})+1, 2)=\" (\")), "
            f"OR(LEFT($C$4, 3)=\"All\", $C$4=\"\", $C$4=B{r}, AND(LEFT($C$4, LEN(B{r}))=B{r}, MID($C$4, LEN(B{r})+1, 2)=\" (\"))"
            f")"
        )
        ws_poly.cell(row=r, column=6, value=f"=IF({formula_filter}, D{r}-E{r}*3, NA())").font = val_font
        ws_poly.cell(row=r, column=7, value=f"=IF({formula_filter}, D{r}-E{r}*6, NA())").font = val_font

        ws_poly.cell(row=r, column=8, value=entry.get("source_paper", "")).font = val_font
        ws_poly.cell(row=r, column=9, value=entry.get("table_name", "")).font = val_font

        if isinstance(c_val, (int, float)) and isinstance(v_val, (int, float)):
            y_vals.append(c_val - v_val * 3)
            y_vals.append(c_val - v_val * 6)

        for c in [1, 2, 8, 9]:
            ws_poly.cell(row=r, column=c).alignment = left_align
            ws_poly.cell(row=r, column=c).border = data_border
        for c in range(3, 8):
            ws_poly.cell(row=r, column=c).alignment = right_align
            ws_poly.cell(row=r, column=c).border = data_border

    # Enable native Excel AutoFilter on the main data table
    ws_poly.auto_filter.ref = f"A{header_row}:I{max(end_data_row, start_data_row)}"

    # Auto-adjust column widths (excluding hidden helper columns)
    for col in ws_poly.columns:
        col_letter = get_column_letter(col[0].column)
        if col_letter in ("J", "L", "M", "N", "O"):
            continue
        vals = [str(cell.value or '') for cell in col if cell.row >= header_row or cell.column in (1, 2, 3)]
        max_len = max(len(v) for v in vals) if vals else 10
        ws_poly.column_dimensions[col_letter].width = max(max_len + 3, 14)

