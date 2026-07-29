import math
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, Reference, Series
from chemstractor.lib.reports.helpers import is_entry_failed, get_entry_errors_info

def build_mark_houwink_sheet(wb, mh_entries: list, font_family: str = "Segoe UI") -> None:
    """Builds the Mark-Houwink worksheet in the Excel workbook."""
    ws_mh = wb.create_sheet(title="Mark-Houwink")
    ws_mh.views.sheetView[0].showGridLines = True
    
    table_header_font = Font(name=font_family, size=10, bold=True, color="FFFFFF")
    table_header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    table_header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    val_font = Font(name=font_family, size=10, color="000000")
    thin_border_side = Side(border_style="thin", color="D9D9D9")
    data_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    left_align = Alignment(horizontal="left", vertical="center")
    right_align = Alignment(horizontal="right", vertical="center")
    
    headers_mh = [
        "Source Paper", "Table", "Polymer (Original)", "Polymer (Clean)", 
        "Solvent (Original)", "Solvent (Clean)", "Temperature (K)", 
        "K Value (mL/g)", "K Transformation", "a Value", "a Transformation", 
        "Errors"
    ]
    
    for col_idx, h in enumerate(headers_mh):
        cell = ws_mh.cell(row=1, column=col_idx + 1, value=h)
        cell.font = table_header_font
        cell.fill = table_header_fill
        cell.alignment = table_header_align
        cell.border = data_border
        
    for val, col in [(3, 13), (6, 14)]:
        cell = ws_mh.cell(row=1, column=col, value=val)
        cell.font = table_header_font
        cell.fill = table_header_fill
        cell.alignment = table_header_align
        cell.border = data_border
        
    mh_counts = {}
    
    for row_idx, entry in enumerate(mh_entries):
        r = row_idx + 2
        ws_mh.cell(row=r, column=1, value=entry.get("source_paper", "")).font = val_font
        ws_mh.cell(row=r, column=2, value=entry.get("table_name", "")).font = val_font
        ws_mh.cell(row=r, column=3, value=entry.get("polymer_name_original", "")).font = val_font
        ws_mh.cell(row=r, column=4, value=entry.get("polymer_name", "")).font = val_font
        ws_mh.cell(row=r, column=5, value=entry.get("solvent_name_original", entry.get("solvent_original", ""))).font = val_font
        ws_mh.cell(row=r, column=6, value=entry.get("solvent_name", entry.get("solvent", ""))).font = val_font
        
        temp = entry.get("temperature_k")
        ws_mh.cell(row=r, column=7, value=float(temp) if isinstance(temp, (int, float)) else temp).font = val_font
        
        k_val = entry.get("K_value")
        ws_mh.cell(row=r, column=8, value=float(k_val) if isinstance(k_val, (int, float)) else k_val).font = val_font
        
        ws_mh.cell(row=r, column=9, value=entry.get("K_transformation", "")).font = val_font
        
        a_val = entry.get("a_value")
        ws_mh.cell(row=r, column=10, value=float(a_val) if isinstance(a_val, (int, float)) else a_val).font = val_font
        
        ws_mh.cell(row=r, column=11, value=entry.get("a_transformation", "")).font = val_font
        
        err_msg, err_fill = get_entry_errors_info(entry)
        err_cell = ws_mh.cell(row=r, column=12, value=err_msg)
        err_cell.font = val_font
        if err_fill:
            err_cell.fill = err_fill
        
        has_failed = is_entry_failed(entry)
        if isinstance(k_val, (int, float)) and k_val > 0 and isinstance(a_val, (int, float)):
            ws_mh.cell(row=r, column=13, value=f"=LOG10(H{r})+J{r}*3").font = val_font
            ws_mh.cell(row=r, column=14, value=f"=LOG10(H{r})+J{r}*6").font = val_font
            
            if not has_failed:
                solv_str = entry.get("solvent_name", entry.get("solvent"))
                pair = (solv_str, entry.get("polymer_name"))
                mh_counts[pair] = mh_counts.get(pair, 0) + 1
            
        for c in [1, 2, 3, 4, 5, 6, 9, 11, 12]:
            cell = ws_mh.cell(row=r, column=c)
            cell.alignment = left_align
            cell.border = data_border
        for c in [7, 8, 10, 13, 14]:
            cell = ws_mh.cell(row=r, column=c)
            cell.alignment = right_align
            cell.border = data_border
            
    # Write Mark-Houwink Summary Table starting at Column P (16)
    ws_mh.cell(row=1, column=16, value="Solvent").font = table_header_font
    ws_mh.cell(row=1, column=16).fill = table_header_fill
    ws_mh.cell(row=1, column=16).alignment = table_header_align
    ws_mh.cell(row=1, column=16).border = data_border
    
    ws_mh.cell(row=1, column=17, value="Polymer").font = table_header_font
    ws_mh.cell(row=1, column=17).fill = table_header_fill
    ws_mh.cell(row=1, column=17).alignment = table_header_align
    ws_mh.cell(row=1, column=17).border = data_border
    
    ws_mh.cell(row=1, column=18, value="Count").font = table_header_font
    ws_mh.cell(row=1, column=18).fill = table_header_fill
    ws_mh.cell(row=1, column=18).alignment = table_header_align
    ws_mh.cell(row=1, column=18).border = data_border
    
    for summary_idx, (pair, count) in enumerate(sorted(mh_counts.items())):
        sr = summary_idx + 2
        ws_mh.cell(row=sr, column=16, value=pair[0]).font = val_font
        ws_mh.cell(row=sr, column=16).alignment = left_align
        ws_mh.cell(row=sr, column=16).border = data_border
        
        ws_mh.cell(row=sr, column=17, value=pair[1]).font = val_font
        ws_mh.cell(row=sr, column=17).alignment = left_align
        ws_mh.cell(row=sr, column=17).border = data_border
        
        ws_mh.cell(row=sr, column=18, value=count).font = val_font
        ws_mh.cell(row=sr, column=18).alignment = right_align
        ws_mh.cell(row=sr, column=18).border = data_border
        
    # Generate and embed Mark-Houwink Line Chart
    if mh_entries:
        try:
            chart_mh = LineChart()
            chart_mh.title = "Mark-Houwink Calibration Curves (log-log Plot)"
            chart_mh.style = 13
            chart_mh.x_axis.title = "log(M / g mol⁻¹)"
            chart_mh.y_axis.title = "log([eta] / mL g⁻¹)"
            chart_mh.x_axis.delete = False
            chart_mh.y_axis.delete = False
            chart_mh.x_axis.tickLblPos = "low"
            chart_mh.y_axis.tickLblPos = "nextTo"
            
            y_vals = []
            for entry in mh_entries:
                K_val = entry.get("K_value")
                a_val = entry.get("a_value")
                has_failed = is_entry_failed(entry)
                if K_val is not None and isinstance(K_val, (int, float)) and K_val > 0 and isinstance(a_val, (int, float)) and not has_failed:
                    log_K = math.log10(K_val)
                    y_vals.append(log_K + a_val * 3)
                    y_vals.append(log_K + a_val * 6)
            if y_vals:
                min_y = min(y_vals)
                max_y = max(y_vals)
                chart_mh.y_axis.scaling.min = float(f"{min_y - 0.5:.2f}")
                chart_mh.y_axis.scaling.max = float(f"{max_y + 0.5:.2f}")
                
            chart_mh.width = 18
            chart_mh.height = 12
            
            for row_idx, entry in enumerate(mh_entries):
                r = row_idx + 2
                K_val = entry.get("K_value")
                a_val = entry.get("a_value")
                has_failed = is_entry_failed(entry)
                if K_val is not None and isinstance(K_val, (int, float)) and K_val > 0 and isinstance(a_val, (int, float)) and not has_failed:
                    series_ref = Reference(ws_mh, min_col=13, max_col=14, min_row=r, max_row=r)
                    solv_str = entry.get('solvent_name', entry.get('solvent'))
                    series = Series(series_ref, title=f"{entry.get('polymer_name')} in {solv_str}")
                    chart_mh.append(series)
                    
            cats_ref = Reference(ws_mh, min_col=13, max_col=14, min_row=1, max_row=1)
            chart_mh.set_categories(cats_ref)
            
            colors = ["1F497D", "C0504D", "9BBB59", "8064A2", "F79646", "4BACC6", "E26B0A", "7030A0", "00B0F0"]
            for s_idx, series in enumerate(chart_mh.series):
                color = colors[s_idx % len(colors)]
                series.graphicalProperties.line = openpyxl.drawing.line.LineProperties(solidFill=color)
                
            ws_mh.add_chart(chart_mh, "S4")
        except Exception:
            pass

    for col in ws_mh.columns:
        vals = [str(cell.value or '') for cell in col]
        max_len = max(len(v) for v in vals) if vals else 10
        col_letter = get_column_letter(col[0].column)
        ws_mh.column_dimensions[col_letter].width = max(max_len + 3, 12)
