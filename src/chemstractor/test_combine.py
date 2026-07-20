import os
import shutil
import json
import unittest
from unittest.mock import MagicMock, patch
# Add src to sys.path so we can import chemstractor
import sys
if hasattr(sys.stdout, 'reconfigure'):
    sys.stdout.reconfigure(encoding='utf-8')
if hasattr(sys.stderr, 'reconfigure'):
    sys.stderr.reconfigure(encoding='utf-8')
if hasattr(sys.__stdout__, 'reconfigure'):
    sys.__stdout__.reconfigure(encoding='utf-8')
if hasattr(sys.__stderr__, 'reconfigure'):
    sys.__stderr__.reconfigure(encoding='utf-8')
sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), '../src')))
from chemstractor.commands.combine import combine_command
from chemstractor.lib.processor import PDFProcessor
from chemstractor.AI import AI, AIPromptResult
class TestCombinePipeline(unittest.TestCase):
    def setUp(self):
        self.test_dir = os.path.abspath(os.path.join(os.path.dirname(__file__), 'runs/mock_combine_run'))
        if os.path.exists(self.test_dir):
            shutil.rmtree(self.test_dir)
        os.makedirs(self.test_dir)
        
        # Paper 1
        self.paper1_dir = os.path.join(self.test_dir, 'test_paper1')
        os.makedirs(os.path.join(self.paper1_dir, 'interpretation'))
        os.makedirs(os.path.join(self.paper1_dir, 'extract/tables/txt'))
        os.makedirs(os.path.join(self.paper1_dir, 'extract/tables/csv'))
        
        # Write dummy output.md to satisfy prepare_processor checks
        with open(os.path.join(self.paper1_dir, 'extract/output.md'), 'w') as f:
            f.write("# Dummy Paper 1")
        with open(os.path.join(self.paper1_dir, 'extract/tables/txt/table1.txt'), 'w') as f:
            f.write("Dummy Table 1")
        with open(os.path.join(self.paper1_dir, 'extract/tables/csv/table1.csv'), 'w') as f:
            f.write("a,b\n1,2")
            
        table1_mh_data = {
            "is_mark_houwink_data": True,
            "mh_entries": [
                {
                    "polymer_name": "PMMA",
                    "solvent": "T0luene",
                    "temperature_k": 298.15,
                    "raw_K_value": 0.012,
                    "K_transformation": "none",
                    "K_value": 0.012,
                    "raw_a_value": 0.7,
                    "a_transformation": "none",
                    "a_value": 0.7
                }
            ]
        }
        with open(os.path.join(self.paper1_dir, 'interpretation/table1_mh.json'), 'w') as f:
            json.dump(table1_mh_data, f)
        with open(os.path.join(self.paper1_dir, 'interpretation/table1_flory.json'), 'w') as f:
            json.dump({"is_flory_data": False, "flory_entries": []}, f)
            
        # Paper 2
        self.paper2_dir = os.path.join(self.test_dir, 'test_paper2')
        os.makedirs(os.path.join(self.paper2_dir, 'interpretation'))
        os.makedirs(os.path.join(self.paper2_dir, 'extract/tables/txt'))
        os.makedirs(os.path.join(self.paper2_dir, 'extract/tables/csv'))
        
        with open(os.path.join(self.paper2_dir, 'extract/output.md'), 'w') as f:
            f.write("# Dummy Paper 2")
        with open(os.path.join(self.paper2_dir, 'extract/tables/txt/table1.txt'), 'w') as f:
            f.write("Dummy Table 1")
        with open(os.path.join(self.paper2_dir, 'extract/tables/csv/table1.csv'), 'w') as f:
            f.write("a,b\n1,2")
            
        table2_flory_data = {
            "is_flory_data": True,
            "flory_entries": [
                {
                    "polymer_name": "Poly(methyl methacrylate)",
                    "solvent": "toluene",
                    "temperature_k": 298.15,
                    "raw_v_value": 0.48,
                    "v_transformation": "none",
                    "v_value": 0.48,
                    "raw_c_value": -7.72,
                    "c_transformation": "none",
                    "c_value": -7.72
                },
                {
                    "polymer_name": "NoiseData",
                    "solvent": "12345",
                    "temperature_k": 298.15,
                    "raw_v_value": 0.5,
                    "v_transformation": "none",
                    "v_value": 0.5,
                    "raw_c_value": -7.5,
                    "c_transformation": "none",
                    "c_value": -7.5
                }
            ]
        }
        with open(os.path.join(self.paper2_dir, 'interpretation/table1_flory.json'), 'w') as f:
            json.dump(table2_flory_data, f)
        with open(os.path.join(self.paper2_dir, 'interpretation/table1_mh.json'), 'w') as f:
            json.dump({"is_mark_houwink_data": False, "mh_entries": []}, f)
    def tearDown(self):
        if os.path.exists(self.test_dir):
            shutil.rmtree(self.test_dir)
    @patch('chemstractor.AI.AI.prompt')
    def test_combine_command(self, mock_prompt):
        # We need mock_prompt to handle calls to match/generate:
        # Prompt 1 (generate PMMA): returns clean_name = "poly(methyl methacrylate)"
        # Prompt 2 (generate T0luene): returns clean_name = "toluene"
        # Prompt 3 (generate NoiseData): returns clean_name = "N/A"
        # Prompt 4 (generate 12345): returns clean_name = "N/A"
        
        # Let's create a list of responses based on the prompt contents
        def side_effect(prompt, schema, **kwargs):
            from pydantic import BaseModel
            
            prompt_str = str(prompt)
            # Check if it's the selection (matching) step or generation step
            # Selection prompt contains "Canonical Options:"
            if "Canonical Options" in prompt_str:
                # Mock selection result (simulate not found)
                class MatchResult(BaseModel):
                    match: str
                return AIPromptResult(success=True, data=MatchResult(match="not found"))
            else:
                # Generation prompt
                class GenResult(BaseModel):
                    clean_name: str
                
                if "Raw chemical name: 'PMMA'" in prompt_str:
                    return AIPromptResult(success=True, data=GenResult(clean_name="poly(methyl methacrylate)"))
                elif "Raw chemical name: 'T0luene'" in prompt_str:
                    return AIPromptResult(success=True, data=GenResult(clean_name="toluene"))
                elif "Raw chemical name: 'NoiseData'" in prompt_str:
                    return AIPromptResult(success=True, data=GenResult(clean_name="N/A"))
                elif "Raw chemical name: '12345'" in prompt_str:
                    return AIPromptResult(success=True, data=GenResult(clean_name="N/A"))
                else:
                    return AIPromptResult(success=True, data=GenResult(clean_name="N/A"))
                    
        mock_prompt.side_effect = side_effect
        
        cache_path = os.path.join(self.test_dir, 'chemical_cache.json')
        output_prefix = os.path.join(self.test_dir, 'combined_output')
        
        # Run command
        combine_command(
            input_dir=self.test_dir,
            output_path=output_prefix,
            cache_path=cache_path
        )
        
        # Assert files were created
        self.assertTrue(os.path.exists(cache_path))
        combine_prefix = os.path.join(output_prefix, 'combined_data')
        self.assertTrue(os.path.exists(combine_prefix + '.json'))
        self.assertTrue(os.path.exists(combine_prefix + '.xlsx'))
        self.assertTrue(os.path.exists(combine_prefix + '_flory_database.pkl'))
        self.assertTrue(os.path.exists(combine_prefix + '_flory_database.csv'))
        self.assertTrue(os.path.exists(combine_prefix + '_mh_database.pkl'))
        self.assertTrue(os.path.exists(combine_prefix + '_mh_database.csv'))
        
        # Validate Cache contents
        with open(cache_path, 'r', encoding='utf-8') as f:
            cache = json.load(f)
            self.assertEqual(cache.get("PMMA"), "poly(methyl methacrylate)")
            self.assertEqual(cache.get("T0luene"), "toluene")
            # Failures (N/A) shouldn't be added to cache or should map to N/A?
            # In our pipeline:
            # if new_name == "N/A": don't save to cache
            self.assertNotIn("NoiseData", cache)
            self.assertNotIn("12345", cache)
            
        # Validate Combined JSON contents
        with open(combine_prefix + '.json', 'r', encoding='utf-8') as f:
            combined = json.load(f)
            
            # Check Mark-Houwink
            mh = combined.get("mark_houwink_entries", [])
            self.assertEqual(len(mh), 1)
            self.assertEqual(mh[0]["polymer_name"], "poly(methyl methacrylate)")
            self.assertEqual(mh[0]["polymer_name_original"], "PMMA")
            self.assertEqual(mh[0]["solvent"], "toluene")
            self.assertEqual(mh[0]["solvent_original"], "T0luene")
            self.assertEqual(mh[0]["failed_fields"], "None")
            
            # Check Flory
            flory = combined.get("flory_entries", [])
            self.assertEqual(len(flory), 2)
            # The first one is Poly(methyl methacrylate) which should resolve to poly(methyl methacrylate)
            self.assertEqual(flory[0]["polymer_name"], "poly(methyl methacrylate)")
            self.assertEqual(flory[0]["solvent"], "toluene")
            self.assertEqual(flory[0]["failed_fields"], "None")
            
            # The second is NoiseData which failed
            self.assertEqual(flory[1]["polymer_name"], "NoiseData")  # remains original raw name on failure
            self.assertEqual(flory[1]["solvent"], "12345")
            self.assertEqual(flory[1]["failed_fields"], "polymer_name, solvent")
            
            # Check Failures
            fails = combined.get("failures", [])
            self.assertEqual(len(fails), 2)
            fields = [x["field"] for x in fails]
            values = [x["value"] for x in fails]
            self.assertIn("polymer_name", fields)
            self.assertIn("solvent", fields)
            self.assertIn("NoiseData", values)
            self.assertIn("12345", values)

        # Validate Excel sheet names and basic content
        import openpyxl
        wb = openpyxl.load_workbook(combine_prefix + '.xlsx')
        sheet_names = wb.sheetnames
        self.assertIn("Mark-Houwink", sheet_names)
        self.assertIn("Flory", sheet_names)
        self.assertIn("Failures", sheet_names)
        self.assertNotIn("Aggregated Flory", sheet_names)
        self.assertNotIn("Aggregated Mark-Houwink", sheet_names)

        # Verify summary tables on sheets (starts at Column P / 16)
        ws_mh = wb["Mark-Houwink"]
        self.assertEqual(ws_mh.cell(row=1, column=16).value, "Solvent")
        self.assertEqual(ws_mh.cell(row=2, column=16).value, "toluene")
        self.assertEqual(ws_mh.cell(row=2, column=17).value, "poly(methyl methacrylate)")
        self.assertEqual(ws_mh.cell(row=2, column=18).value, 1)

        ws_flory = wb["Flory"]
        self.assertEqual(ws_flory.cell(row=1, column=16).value, "Solvent")
        self.assertEqual(ws_flory.cell(row=2, column=16).value, "toluene")
        self.assertEqual(ws_flory.cell(row=2, column=17).value, "poly(methyl methacrylate)")
        self.assertEqual(ws_flory.cell(row=2, column=18).value, 1)

        # Verify computed plot points in columns M (13) and N (14)
        self.assertEqual(ws_flory.cell(row=1, column=13).value, 3)
        self.assertEqual(ws_flory.cell(row=1, column=14).value, 6)
        self.assertEqual(ws_flory.cell(row=2, column=13).value, "=H2-J2*3")
        self.assertEqual(ws_flory.cell(row=2, column=14).value, "=H2-J2*6")

        self.assertEqual(ws_mh.cell(row=1, column=13).value, 3)
        self.assertEqual(ws_mh.cell(row=1, column=14).value, 6)
        self.assertEqual(ws_mh.cell(row=2, column=13).value, "=LOG10(H2)+J2*3")
        self.assertEqual(ws_mh.cell(row=2, column=14).value, "=LOG10(H2)+J2*6")

        # Verify LineCharts are present in both sheets
        self.assertEqual(len(ws_mh._charts), 1)
        self.assertEqual(len(ws_flory._charts), 1)

        # Validate Databases (Pandas DataFrames are flat)
        import pandas as pd
        
        # 1. Flory
        df_flory = pd.read_pickle(combine_prefix + '_flory_database.pkl')
        self.assertEqual(len(df_flory), 1)
        self.assertEqual(df_flory.iloc[0]["solvent"], "toluene")
        self.assertEqual(df_flory.iloc[0]["polymer"], "poly(methyl methacrylate)")
        self.assertEqual(df_flory.iloc[0]["c_value"], -7.72)
        self.assertEqual(df_flory.iloc[0]["v_value"], 0.48)

        # 2. Mark-Houwink
        df_mh = pd.read_pickle(combine_prefix + '_mh_database.pkl')
        self.assertEqual(len(df_mh), 1)
        self.assertEqual(df_mh.iloc[0]["solvent"], "toluene")
        self.assertEqual(df_mh.iloc[0]["polymer"], "poly(methyl methacrylate)")
        self.assertEqual(df_mh.iloc[0]["K_value"], 0.012)
        self.assertEqual(df_mh.iloc[0]["a_value"], 0.7)
        
if __name__ == '__main__':
    unittest.main()
