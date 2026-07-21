import os
import sys
import json
import time
from datetime import datetime
import openpyxl
import pandas as pd
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from rich.console import Console
from rich.table import Table
from rich.rule import Rule
from rich.live import Live

from chemstractor.lib.processor import PDFProcessor
from chemstractor.AI import AI, pricing_matrix
from chemstractor.lib.combiner import gather_and_homogenise

import os
import sys
import json
import time
import openpyxl
import pandas as pd
import math
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, Reference
from rich.console import Console
from rich.table import Table
from rich.rule import Rule
from rich.live import Live
from rich.progress import Progress, SpinnerColumn, BarColumn, TextColumn, TaskProgressColumn, TimeElapsedColumn, MofNCompleteColumn
from rich.tree import Tree
from rich.console import Group

from chemstractor.lib.processor import PDFProcessor
from chemstractor.AI import AI, pricing_matrix
from chemstractor.lib.combiner import gather_and_homogenise

from openpyxl.chart import LineChart, Reference, Series


def create_combined_excel(dest_path: str, combined_data: dict) -> None:
    """Generates a nicely formatted combined Excel report containing Summary, Papers, Polymers, Flory, Mark-Houwink, and Failures sheets."""
    wb = openpyxl.Workbook()
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    font_family = "Segoe UI"
    
    # Styles
    table_header_font = Font(name=font_family, size=10, bold=True, color="FFFFFF")
    table_header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    table_header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    section_title_font = Font(name=font_family, size=11, bold=True, color="1F497D")
    val_font = Font(name=font_family, size=10, color="000000")
    thin_border_side = Side(border_style="thin", color="D9D9D9")
    data_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    
    fail_row_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
    
    left_align = Alignment(horizontal="left", vertical="center")
    right_align = Alignment(horizontal="right", vertical="center")
    center_align = Alignment(horizontal="center", vertical="center")
    
    flory_entries = combined_data.get("flory_entries", [])
    mh_entries = combined_data.get("mark_houwink_entries", [])
    failures = combined_data.get("failures", [])
    papers_summary = combined_data.get("papers_summary", [])
    
    # Pre-calculate failure stats for Summary sheet
    failure_reasons_count = {}
    total_mh_failed = 0
    for entry in mh_entries:
        ff = entry.get("failed_fields", "None")
        poly = entry.get("polymer_name", "")
        solv = entry.get("solvent", "")
        if ff != "None" or poly == "N/A" or solv == "N/A":
            total_mh_failed += 1
            reason_str = ff if ff != "None" else "Unresolved chemical name (N/A)"
            failure_reasons_count[reason_str] = failure_reasons_count.get(reason_str, 0) + 1

    total_flory_failed = 0
    for entry in flory_entries:
        ff = entry.get("failed_fields", "None")
        poly = entry.get("polymer_name", "")
        solv = entry.get("solvent", "")
        if ff != "None" or poly == "N/A" or solv == "N/A":
            total_flory_failed += 1
            reason_str = ff if ff != "None" else "Unresolved chemical name (N/A)"
            failure_reasons_count[reason_str] = failure_reasons_count.get(reason_str, 0) + 1

    for fail in failures:
        r_str = fail.get("reason", "Unknown failure")
        failure_reasons_count[r_str] = failure_reasons_count.get(r_str, 0) + 1

    total_collected = len(flory_entries) + len(mh_entries)
    total_failures_all = total_flory_failed + total_mh_failed + len(failures)

    # ---------------------------------------------------------
    # 1. Summary Sheet
    # ---------------------------------------------------------
    ws_sum = wb.create_sheet(title="Summary")
    ws_sum.views.sheetView[0].showGridLines = True

    ws_sum.cell(row=1, column=1, value="Extraction & Homogenisation Statistics").font = section_title_font
    
    headers_sum_kpi = ["Metric", "Value"]
    for c_idx, h in enumerate(headers_sum_kpi):
        cell = ws_sum.cell(row=2, column=c_idx + 1, value=h)
        cell.font = table_header_font
        cell.fill = table_header_fill
        cell.alignment = table_header_align
        cell.border = data_border

    success_rate = ((total_collected - total_failures_all) / total_collected * 100) if total_collected > 0 else 0.0

    kpis = [
        ("Total Data Points Collected", total_collected),
        ("Flory Data Points", len(flory_entries)),
        ("Mark-Houwink Data Points", len(mh_entries)),
        ("Total Failures / Issues", total_failures_all),
        ("Success Rate (%)", f"{success_rate:.1f}%")
    ]

    for r_idx, (m, v) in enumerate(kpis):
        r = r_idx + 3
        c1 = ws_sum.cell(row=r, column=1, value=m)
        c1.font = val_font
        c1.alignment = left_align
        c1.border = data_border
        
        c2 = ws_sum.cell(row=r, column=2, value=v)
        c2.font = val_font
        c2.alignment = right_align
        c2.border = data_border

    # Failure Reasons breakdown table
    ws_sum.cell(row=10, column=1, value="Failure Reasons Breakdown").font = section_title_font

    headers_sum_fail = ["Failure Reason / Field", "Count"]
    for c_idx, h in enumerate(headers_sum_fail):
        cell = ws_sum.cell(row=11, column=c_idx + 1, value=h)
        cell.font = table_header_font
        cell.fill = table_header_fill
        cell.alignment = table_header_align
        cell.border = data_border

    if failure_reasons_count:
        sorted_reasons = sorted(failure_reasons_count.items(), key=lambda x: x[1], reverse=True)
        for r_idx, (reason, count) in enumerate(sorted_reasons):
            r = r_idx + 12
            c1 = ws_sum.cell(row=r, column=1, value=reason)
            c1.font = val_font
            c1.alignment = left_align
            c1.border = data_border

            c2 = ws_sum.cell(row=r, column=2, value=count)
            c2.font = val_font
            c2.alignment = right_align
            c2.border = data_border
    else:
        c1 = ws_sum.cell(row=12, column=1, value="None (All extractions succeeded)")
        c1.font = val_font
        c1.alignment = left_align
        c1.border = data_border
        c2 = ws_sum.cell(row=12, column=2, value=0)
        c2.font = val_font
        c2.alignment = right_align
        c2.border = data_border

    for col in ws_sum.columns:
        vals = [str(cell.value or '') for cell in col]
        max_len = max(len(v) for v in vals) if vals else 10
        col_letter = get_column_letter(col[0].column)
        ws_sum.column_dimensions[col_letter].width = max(max_len + 4, 15)

    # ---------------------------------------------------------
    # 2. Papers Sheet
    # ---------------------------------------------------------
    ws_papers = wb.create_sheet(title="Papers")
    ws_papers.views.sheetView[0].showGridLines = True

    headers_papers = [
        "Source Paper", "Total Tables", "Selected Tables", 
        "Flory Data Points", "Mark-Houwink Data Points", "Failed Data Points"
    ]
    for col_idx, h in enumerate(headers_papers):
        cell = ws_papers.cell(row=1, column=col_idx + 1, value=h)
        cell.font = table_header_font
        cell.fill = table_header_fill
        cell.alignment = table_header_align
        cell.border = data_border

    for row_idx, p_info in enumerate(papers_summary):
        r = row_idx + 2
        ws_papers.cell(row=r, column=1, value=p_info.get("source_paper", "")).font = val_font
        ws_papers.cell(row=r, column=2, value=p_info.get("total_tables", 0)).font = val_font
        ws_papers.cell(row=r, column=3, value=p_info.get("selected_tables", 0)).font = val_font
        ws_papers.cell(row=r, column=4, value=p_info.get("flory_count", 0)).font = val_font
        ws_papers.cell(row=r, column=5, value=p_info.get("mh_count", 0)).font = val_font
        ws_papers.cell(row=r, column=6, value=p_info.get("failed_count", 0)).font = val_font

        ws_papers.cell(row=r, column=1).alignment = left_align
        ws_papers.cell(row=r, column=1).border = data_border
        for c in range(2, 7):
            ws_papers.cell(row=r, column=c).alignment = right_align
            ws_papers.cell(row=r, column=c).border = data_border

    for col in ws_papers.columns:
        vals = [str(cell.value or '') for cell in col]
        max_len = max(len(v) for v in vals) if vals else 10
        col_letter = get_column_letter(col[0].column)
        ws_papers.column_dimensions[col_letter].width = max(max_len + 3, 14)

    # ---------------------------------------------------------
    # 3. Polymers Sheet (Polymer-centric Flory data)
    # ---------------------------------------------------------
    ws_poly = wb.create_sheet(title="Polymers")
    ws_poly.views.sheetView[0].showGridLines = True

    headers_poly = [
        "Polymer (Clean)", "Solvent (Clean)", "Temperature (K)", 
        "c Value (log Constant)", "v Value (Scaling Exponent)", "3", "6"
    ]
    for col_idx, h in enumerate(headers_poly):
        cell = ws_poly.cell(row=1, column=col_idx + 1, value=h)
        cell.font = table_header_font
        cell.fill = table_header_fill
        cell.alignment = table_header_align
        cell.border = data_border

    # Filter out failed entries
    valid_flory_poly = []
    for entry in flory_entries:
        c_val = entry.get("c_value")
        v_val = entry.get("v_value")
        has_failed = entry.get("failed_fields", "None") != "None" or entry.get("polymer_name") == "N/A" or entry.get("solvent") == "N/A"
        if isinstance(c_val, (int, float)) and isinstance(v_val, (int, float)) and not has_failed:
            valid_flory_poly.append(entry)

    # Sort Flory entries polymer-centric
    sorted_flory_poly = sorted(
        valid_flory_poly, 
        key=lambda x: (
            (x.get("polymer_name") or "").lower(),
            (x.get("solvent") or "").lower()
        )
    )

    for row_idx, entry in enumerate(sorted_flory_poly):
        r = row_idx + 2
        ws_poly.cell(row=r, column=1, value=entry.get("polymer_name", "")).font = val_font
        ws_poly.cell(row=r, column=2, value=entry.get("solvent", "")).font = val_font

        temp = entry.get("temperature_k")
        ws_poly.cell(row=r, column=3, value=float(temp) if isinstance(temp, (int, float)) else temp).font = val_font

        c_val = entry.get("c_value")
        ws_poly.cell(row=r, column=4, value=float(c_val) if isinstance(c_val, (int, float)) else c_val).font = val_font

        v_val = entry.get("v_value")
        ws_poly.cell(row=r, column=5, value=float(v_val) if isinstance(v_val, (int, float)) else v_val).font = val_font

        ws_poly.cell(row=r, column=6, value=f"=D{r}-E{r}*3").font = val_font
        ws_poly.cell(row=r, column=7, value=f"=D{r}-E{r}*6").font = val_font

        for c in [1, 2]:
            ws_poly.cell(row=r, column=c).alignment = left_align
            ws_poly.cell(row=r, column=c).border = data_border
        for c in range(3, 8):
            ws_poly.cell(row=r, column=c).alignment = right_align
            ws_poly.cell(row=r, column=c).border = data_border

    for col in ws_poly.columns:
        vals = [str(cell.value or '') for cell in col]
        max_len = max(len(v) for v in vals) if vals else 10
        col_letter = get_column_letter(col[0].column)
        ws_poly.column_dimensions[col_letter].width = max(max_len + 3, 12)

    # ---------------------------------------------------------
    # 4. Flory Sheet
    # ---------------------------------------------------------
    ws_flory = wb.create_sheet(title="Flory")
    ws_flory.views.sheetView[0].showGridLines = True
    
    headers_flory = [
        "Source Paper", "Table", "Polymer (Original)", "Polymer (Clean)", 
        "Solvent (Original)", "Solvent (Clean)", "Temperature (K)", 
        "c Value (log Constant)", "c Transformation", "v Value (Scaling Exponent)", "v Transformation", "Failed Field"
    ]
    
    for col_idx, h in enumerate(headers_flory):
        cell = ws_flory.cell(row=1, column=col_idx + 1, value=h)
        cell.font = table_header_font
        cell.fill = table_header_fill
        cell.alignment = table_header_align
        cell.border = data_border
        
    for val, col in [(3, 13), (6, 14)]:
        cell = ws_flory.cell(row=1, column=col, value=val)
        cell.font = table_header_font
        cell.fill = table_header_fill
        cell.alignment = table_header_align
        cell.border = data_border
        
    flory_counts = {}
    
    for row_idx, entry in enumerate(flory_entries):
        r = row_idx + 2
        ws_flory.cell(row=r, column=1, value=entry.get("source_paper", "")).font = val_font
        ws_flory.cell(row=r, column=2, value=entry.get("table_name", "")).font = val_font
        ws_flory.cell(row=r, column=3, value=entry.get("polymer_name_original", "")).font = val_font
        ws_flory.cell(row=r, column=4, value=entry.get("polymer_name", "")).font = val_font
        ws_flory.cell(row=r, column=5, value=entry.get("solvent_original", "")).font = val_font
        ws_flory.cell(row=r, column=6, value=entry.get("solvent", "")).font = val_font
        
        temp = entry.get("temperature_k")
        ws_flory.cell(row=r, column=7, value=float(temp) if isinstance(temp, (int, float)) else temp).font = val_font
        
        c_val = entry.get("c_value")
        ws_flory.cell(row=r, column=8, value=float(c_val) if isinstance(c_val, (int, float)) else c_val).font = val_font
        
        ws_flory.cell(row=r, column=9, value=entry.get("c_transformation", "")).font = val_font
        
        v_val = entry.get("v_value")
        ws_flory.cell(row=r, column=10, value=float(v_val) if isinstance(v_val, (int, float)) else v_val).font = val_font
        
        ws_flory.cell(row=r, column=11, value=entry.get("v_transformation", "")).font = val_font
        ws_flory.cell(row=r, column=12, value=entry.get("failed_fields", "None")).font = val_font
        
        has_failed = entry.get("failed_fields", "None") != "None" or entry.get("polymer_name") == "N/A" or entry.get("solvent") == "N/A"
        if c_val is not None and v_val is not None and not has_failed:
            ws_flory.cell(row=r, column=13, value=f"=H{r}-J{r}*3").font = val_font
            ws_flory.cell(row=r, column=14, value=f"=H{r}-J{r}*6").font = val_font
            
            pair = (entry.get("solvent"), entry.get("polymer_name"))
            flory_counts[pair] = flory_counts.get(pair, 0) + 1
            
        for c in [1, 2, 3, 4, 5, 6, 9, 11, 12]:
            cell = ws_flory.cell(row=r, column=c)
            cell.alignment = left_align
            cell.border = data_border
            if has_failed:
                cell.fill = fail_row_fill
        for c in [7, 8, 10, 13, 14]:
            cell = ws_flory.cell(row=r, column=c)
            cell.alignment = right_align
            cell.border = data_border
            if has_failed:
                cell.fill = fail_row_fill
            
    # Write Flory Summary Table starting at Column P (16)
    ws_flory.cell(row=1, column=16, value="Solvent").font = table_header_font
    ws_flory.cell(row=1, column=16).fill = table_header_fill
    ws_flory.cell(row=1, column=16).alignment = table_header_align
    ws_flory.cell(row=1, column=16).border = data_border
    
    ws_flory.cell(row=1, column=17, value="Polymer").font = table_header_font
    ws_flory.cell(row=1, column=17).fill = table_header_fill
    ws_flory.cell(row=1, column=17).alignment = table_header_align
    ws_flory.cell(row=1, column=17).border = data_border
    
    ws_flory.cell(row=1, column=18, value="Count").font = table_header_font
    ws_flory.cell(row=1, column=18).fill = table_header_fill
    ws_flory.cell(row=1, column=18).alignment = table_header_align
    ws_flory.cell(row=1, column=18).border = data_border
    
    for summary_idx, (pair, count) in enumerate(sorted(flory_counts.items())):
        sr = summary_idx + 2
        ws_flory.cell(row=sr, column=16, value=pair[0]).font = val_font
        ws_flory.cell(row=sr, column=16).alignment = left_align
        ws_flory.cell(row=sr, column=16).border = data_border
        
        ws_flory.cell(row=sr, column=17, value=pair[1]).font = val_font
        ws_flory.cell(row=sr, column=17).alignment = left_align
        ws_flory.cell(row=sr, column=17).border = data_border
        
        ws_flory.cell(row=sr, column=18, value=count).font = val_font
        ws_flory.cell(row=sr, column=18).alignment = right_align
        ws_flory.cell(row=sr, column=18).border = data_border
        
    # Generate and embed Flory Line Chart
    if flory_entries:
        try:
            chart_flory = LineChart()
            chart_flory.title = "Flory Calibration Curves (log-log Plot)"
            chart_flory.style = 13
            chart_flory.x_axis.title = "log(M / g mol⁻¹)"
            chart_flory.y_axis.title = "log(D / m² s⁻¹)"
            chart_flory.x_axis.delete = False
            chart_flory.y_axis.delete = False
            chart_flory.x_axis.tickLblPos = "low"
            chart_flory.y_axis.tickLblPos = "nextTo"
            
            y_vals = []
            for entry in flory_entries:
                c_val = entry.get("c_value")
                v_val = entry.get("v_value")
                has_failed = entry.get("failed_fields", "None") != "None" or entry.get("polymer_name") == "N/A" or entry.get("solvent") == "N/A"
                if c_val is not None and v_val is not None and not has_failed:
                    y_vals.append(c_val - v_val * 3)
                    y_vals.append(c_val - v_val * 6)
            if y_vals:
                min_y = min(y_vals)
                max_y = max(y_vals)
                chart_flory.y_axis.scaling.min = float(f"{min_y - 0.5:.2f}")
                chart_flory.y_axis.scaling.max = float(f"{max_y + 0.5:.2f}")
                
            chart_flory.width = 18
            chart_flory.height = 12
            
            for row_idx, entry in enumerate(flory_entries):
                r = row_idx + 2
                c_val = entry.get("c_value")
                v_val = entry.get("v_value")
                has_failed = entry.get("failed_fields", "None") != "None" or entry.get("polymer_name") == "N/A" or entry.get("solvent") == "N/A"
                if c_val is not None and v_val is not None and not has_failed:
                    series_ref = Reference(ws_flory, min_col=13, max_col=14, min_row=r, max_row=r)
                    series = Series(series_ref, title=f"{entry.get('polymer_name')} in {entry.get('solvent')}")
                    chart_flory.append(series)
                    
            cats_ref = Reference(ws_flory, min_col=13, max_col=14, min_row=1, max_row=1)
            chart_flory.set_categories(cats_ref)
            
            colors = ["1F497D", "C0504D", "9BBB59", "8064A2", "F79646", "4BACC6", "E26B0A", "7030A0", "00B0F0"]
            for s_idx, series in enumerate(chart_flory.series):
                color = colors[s_idx % len(colors)]
                series.graphicalProperties.line = openpyxl.drawing.line.LineProperties(solidFill=color)
                
            ws_flory.add_chart(chart_flory, "T4")
        except Exception:
            pass
            
    for col in ws_flory.columns:
        vals = [str(cell.value or '') for cell in col]
        max_len = max(len(v) for v in vals) if vals else 10
        col_letter = get_column_letter(col[0].column)
        ws_flory.column_dimensions[col_letter].width = max(max_len + 3, 12)

    # ---------------------------------------------------------
    # 5. Mark-Houwink Sheet
    # ---------------------------------------------------------
    ws_mh = wb.create_sheet(title="Mark-Houwink")
    ws_mh.views.sheetView[0].showGridLines = True
    
    headers_mh = [
        "Source Paper", "Table", "Polymer (Original)", "Polymer (Clean)", 
        "Solvent (Original)", "Solvent (Clean)", "Temperature (K)", 
        "K Value (mL/g)", "K Transformation", "a Value", "a Transformation", "Failed Field"
    ]
    
    for col_idx, h in enumerate(headers_mh):
        cell = ws_mh.cell(row=1, column=col_idx + 1, value=h)
        cell.font = table_header_font
        cell.fill = table_header_fill
        cell.alignment = table_header_align
        cell.border = data_border
        
    for val, col in [(3, 13), (6, 14)]:
        cell = ws_mh.cell(row=1, column=col, value=val)
        cell.font = table_header_font
        cell.fill = table_header_fill
        cell.alignment = table_header_align
        cell.border = data_border
        
    mh_counts = {}
    
    for row_idx, entry in enumerate(mh_entries):
        r = row_idx + 2
        ws_mh.cell(row=r, column=1, value=entry.get("source_paper", "")).font = val_font
        ws_mh.cell(row=r, column=2, value=entry.get("table_name", "")).font = val_font
        ws_mh.cell(row=r, column=3, value=entry.get("polymer_name_original", "")).font = val_font
        ws_mh.cell(row=r, column=4, value=entry.get("polymer_name", "")).font = val_font
        ws_mh.cell(row=r, column=5, value=entry.get("solvent_original", "")).font = val_font
        ws_mh.cell(row=r, column=6, value=entry.get("solvent", "")).font = val_font
        
        temp = entry.get("temperature_k")
        ws_mh.cell(row=r, column=7, value=float(temp) if temp is not None else None).font = val_font
        
        k_val = entry.get("K_value")
        ws_mh.cell(row=r, column=8, value=float(k_val) if k_val is not None else None).font = val_font
        
        ws_mh.cell(row=r, column=9, value=entry.get("K_transformation", "")).font = val_font
        
        a_val = entry.get("a_value")
        ws_mh.cell(row=r, column=10, value=float(a_val) if a_val is not None else None).font = val_font
        
        ws_mh.cell(row=r, column=11, value=entry.get("a_transformation", "")).font = val_font
        ws_mh.cell(row=r, column=12, value=entry.get("failed_fields", "None")).font = val_font
        
        has_failed = entry.get("failed_fields", "None") != "None" or entry.get("polymer_name") == "N/A" or entry.get("solvent") == "N/A"
        if k_val is not None and k_val > 0 and a_val is not None and not has_failed:
            ws_mh.cell(row=r, column=13, value=f"=LOG10(H{r})+J{r}*3").font = val_font
            ws_mh.cell(row=r, column=14, value=f"=LOG10(H{r})+J{r}*6").font = val_font
            
            pair = (entry.get("solvent"), entry.get("polymer_name"))
            mh_counts[pair] = mh_counts.get(pair, 0) + 1
            
        for c in [1, 2, 3, 4, 5, 6, 9, 11, 12]:
            cell = ws_mh.cell(row=r, column=c)
            cell.alignment = left_align
            cell.border = data_border
            if has_failed:
                cell.fill = fail_row_fill
        for c in [7, 8, 10, 13, 14]:
            cell = ws_mh.cell(row=r, column=c)
            cell.alignment = right_align
            cell.border = data_border
            if has_failed:
                cell.fill = fail_row_fill
            
    # Write Mark-Houwink Summary Table starting at Column P (16)
    ws_mh.cell(row=1, column=16, value="Solvent").font = table_header_font
    ws_mh.cell(row=1, column=16).fill = table_header_fill
    ws_mh.cell(row=1, column=16).alignment = table_header_align
    ws_mh.cell(row=1, column=16).border = data_border
    
    ws_mh.cell(row=1, column=17, value="Polymer").font = table_header_font
    ws_mh.cell(row=1, column=17).fill = table_header_fill
    ws_mh.cell(row=1, column=17).alignment = table_header_align
    ws_mh.cell(row=1, column=17).border = data_border
    
    ws_mh.cell(row=1, column=18, value="Count").font = table_header_font
    ws_mh.cell(row=1, column=18).fill = table_header_fill
    ws_mh.cell(row=1, column=18).alignment = table_header_align
    ws_mh.cell(row=1, column=18).border = data_border
    
    for summary_idx, (pair, count) in enumerate(sorted(mh_counts.items())):
        sr = summary_idx + 2
        ws_mh.cell(row=sr, column=16, value=pair[0]).font = val_font
        ws_mh.cell(row=sr, column=16).alignment = left_align
        ws_mh.cell(row=sr, column=16).border = data_border
        
        ws_mh.cell(row=sr, column=17, value=pair[1]).font = val_font
        ws_mh.cell(row=sr, column=17).alignment = left_align
        ws_mh.cell(row=sr, column=17).border = data_border
        
        ws_mh.cell(row=sr, column=18, value=count).font = val_font
        ws_mh.cell(row=sr, column=18).alignment = right_align
        ws_mh.cell(row=sr, column=18).border = data_border
        
    # Generate and embed Mark-Houwink Line Chart
    if mh_entries:
        try:
            chart_mh = LineChart()
            chart_mh.title = "Mark-Houwink Calibration Curves (log-log Plot)"
            chart_mh.style = 13
            chart_mh.x_axis.title = "log(M / g mol⁻¹)"
            chart_mh.y_axis.title = "log([eta] / mL g⁻¹)"
            chart_mh.x_axis.delete = False
            chart_mh.y_axis.delete = False
            chart_mh.x_axis.tickLblPos = "low"
            chart_mh.y_axis.tickLblPos = "nextTo"
            
            y_vals = []
            for entry in mh_entries:
                K_val = entry.get("K_value")
                a_val = entry.get("a_value")
                has_failed = entry.get("failed_fields", "None") != "None" or entry.get("polymer_name") == "N/A" or entry.get("solvent") == "N/A"
                if K_val is not None and K_val > 0 and a_val is not None and not has_failed:
                    log_K = math.log10(K_val)
                    y_vals.append(log_K + a_val * 3)
                    y_vals.append(log_K + a_val * 6)
            if y_vals:
                min_y = min(y_vals)
                max_y = max(y_vals)
                chart_mh.y_axis.scaling.min = float(f"{min_y - 0.5:.2f}")
                chart_mh.y_axis.scaling.max = float(f"{max_y + 0.5:.2f}")
                
            chart_mh.width = 18
            chart_mh.height = 12
            
            for row_idx, entry in enumerate(mh_entries):
                r = row_idx + 2
                K_val = entry.get("K_value")
                a_val = entry.get("a_value")
                has_failed = entry.get("failed_fields", "None") != "None" or entry.get("polymer_name") == "N/A" or entry.get("solvent") == "N/A"
                if K_val is not None and K_val > 0 and a_val is not None and not has_failed:
                    series_ref = Reference(ws_mh, min_col=13, max_col=14, min_row=r, max_row=r)
                    series = Series(series_ref, title=f"{entry.get('polymer_name')} in {entry.get('solvent')}")
                    chart_mh.append(series)
                    
            cats_ref = Reference(ws_mh, min_col=13, max_col=14, min_row=1, max_row=1)
            chart_mh.set_categories(cats_ref)
            
            colors = ["1F497D", "C0504D", "9BBB59", "8064A2", "F79646", "4BACC6", "E26B0A", "7030A0", "00B0F0"]
            for s_idx, series in enumerate(chart_mh.series):
                color = colors[s_idx % len(colors)]
                series.graphicalProperties.line = openpyxl.drawing.line.LineProperties(solidFill=color)
                
            ws_mh.add_chart(chart_mh, "T4")
        except Exception:
            pass

    for col in ws_mh.columns:
        vals = [str(cell.value or '') for cell in col]
        max_len = max(len(v) for v in vals) if vals else 10
        col_letter = get_column_letter(col[0].column)
        ws_mh.column_dimensions[col_letter].width = max(max_len + 3, 12)

    # ---------------------------------------------------------
    # 6. Failures Sheet
    # ---------------------------------------------------------
    ws_fail = wb.create_sheet(title="Failures")
    ws_fail.views.sheetView[0].showGridLines = True
    
    headers_fail = ["Source Paper", "Table", "Field Name", "Raw Value", "Failure Reason"]
    for col_idx, h in enumerate(headers_fail):
        cell = ws_fail.cell(row=1, column=col_idx + 1, value=h)
        cell.font = table_header_font
        cell.fill = table_header_fill
        cell.alignment = table_header_align
        cell.border = data_border
        
    for row_idx, entry in enumerate(failures):
        r = row_idx + 2
        ws_fail.cell(row=r, column=1, value=entry.get("source_paper", "")).font = val_font
        ws_fail.cell(row=r, column=2, value=entry.get("table", "")).font = val_font
        ws_fail.cell(row=r, column=3, value=entry.get("field", "")).font = val_font
        ws_fail.cell(row=r, column=4, value=entry.get("value", "")).font = val_font
        ws_fail.cell(row=r, column=5, value=entry.get("reason", "")).font = val_font
        
        for c in range(1, 6):
            cell = ws_fail.cell(row=r, column=c)
            cell.alignment = left_align
            cell.border = data_border
            cell.fill = fail_row_fill
            
    for col in ws_fail.columns:
        vals = [str(cell.value or '') for cell in col]
        max_len = max(len(v) for v in vals) if vals else 10
        col_letter = get_column_letter(col[0].column)
        ws_fail.column_dimensions[col_letter].width = max(max_len + 3, 12)

    for ws in wb.worksheets:
        ws.freeze_panes = "A2"

    wb.save(dest_path)


def combine_command(input_dir: str, output_path: str = None, cache_path: str = None) -> None:
    """Executes the combine & homogenisation process across all process output folders in input_dir."""
    console = Console(file=sys.__stdout__)
    
    if not os.path.exists(input_dir):
        console.print(f"[bold red]Error: Input directory '{input_dir}' does not exist.[/bold red]")
        sys.exit(1)
        
    if not os.path.isdir(input_dir):
        console.print(f"[bold red]Error: '{input_dir}' is not a directory.[/bold red]")
        sys.exit(1)
        
    # Default cache path to input_dir/chemical_cache.json if not specified
    if cache_path is None:
        cache_path = os.path.join(input_dir, "chemical_cache.json")
        
    dateTime = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        
    # Default output_path to input_dir/combined_data_{dateTime} if not specified
    if output_path is None:
        output_prefix = os.path.join(input_dir, f"combined_data_{dateTime}")
    else:
        # Check if it has extension or represents a folder
        if output_path.endswith(".json") or output_path.endswith(".xlsx"):
            output_prefix = f"{os.path.splitext(output_path)[0]}_{dateTime}"
        else:
            output_prefix = os.path.join(output_path, f"combined_data_{dateTime}")
            
    # Resolve exact file paths
    json_out = f"{output_prefix}.json"
    xlsx_out = f"{output_prefix}.xlsx"
    
    # 1. Grab all subdirectories in input_dir
    subfolders = [
        os.path.join(input_dir, d) for d in os.listdir(input_dir)
        if os.path.isdir(os.path.join(input_dir, d))
    ]
    
    # 2. Filter directories that look like process outputs
    valid_subfolders = []
    for sf in subfolders:
        extract_dir = os.path.join(sf, "extract")
        interp_dir = os.path.join(sf, "interpretation")
        if os.path.isdir(extract_dir) or os.path.isdir(interp_dir):
            valid_subfolders.append(sf)
            
    if not valid_subfolders:
        console.print("[bold red]Error: No subfolders containing extracted or interpreted data found in the input directory.[/bold red]")
        sys.exit(1)
        
    console.print(Rule(title=f"[bold magenta]COMBINING & HOMOGENISING DATA ({len(valid_subfolders)} PAPERS)[/bold magenta]", style="magenta"))
    console.print(f"Inputs dir: [cyan]{input_dir}[/cyan]")
    console.print(f"Cache file: [cyan]{cache_path}[/cyan]")
    console.print()
    
    # 3. Load PDFProcessor instances
    processors = []
    with console.status("[bold green]Loading process outputs into PDFProcessors...", spinner="dots"):
        for sf in sorted(valid_subfolders):
            try:
                proc = PDFProcessor()
                proc.load_output(sf)
                has_interp = (
                    any(item is not None for item in proc.interpretation_flory_data_list) or
                    any(item is not None for item in proc.interpretation_mh_data_list)
                )
                if has_interp:
                    processors.append(proc)
            except Exception as e:
                console.print(f"[yellow]⚠[/yellow] Warning: Failed to load processor output from '{sf}': {e}")
                
    if not processors:
        console.print("[bold yellow]No papers with valid interpretation data were found. Nothing to combine.[/bold yellow]")
        return
        
    console.print(f"Loaded [green]{len(processors)}[/green] papers containing interpreted table data.")
    
    # 4. Run homogenisation data pipeline
    start_time = time.time()
    ai_instance = AI.get_instance()
    selected_model = ai_instance.selected_model

    
    combine_tree = Tree(f"[bold magenta]Homogenising Chemical Names using {selected_model}[/bold magenta]")
    progress = Progress(
        SpinnerColumn(),
        TextColumn("[bold cyan]{task.description}[/bold cyan]"),
        BarColumn(),
        TaskProgressColumn(),
        MofNCompleteColumn(),
        TimeElapsedColumn(),
        console=console
    )
    task_id = progress.add_task("Homogenising papers...", total=len(processors))
    group = Group(progress, combine_tree)
    
    results = None
    with Live(group, console=console, auto_refresh=True, refresh_per_second=12):
        for event in gather_and_homogenise(processors, cache_path, ai_instance):
            status = event.get("status")
            if status == "paper_start":
                paper_name = event["paper_name"]
                progress.update(task_id, description=f"Homogenising [cyan]{paper_name}[/cyan]...")
            elif status == "paper_complete":
                paper_name = event["paper_name"]
                mh_cnt = event.get("mh_count", 0)
                flory_cnt = event.get("flory_count", 0)
                combine_tree.add(
                    f"[green]✓[/green] [bold]{paper_name}[/bold] "
                    f"[dim]({flory_cnt} Flory, {mh_cnt} Mark-Houwink)[/dim]"
                )
                progress.advance(task_id)
            elif status == "complete":
                results = event["results"]
                progress.update(task_id, description="[bold green]Homogenisation complete![/bold green]")
                
    elapsed = time.time() - start_time
    
    # Build flat (unaggregated) pandas dataframes and export
    # 1. Flory flat database
    flory_entries = results.get("flory_entries", [])
    valid_flory = []
    for entry in flory_entries:
        if entry.get("failed_fields") and entry["failed_fields"] != "None":
            continue
        poly = entry.get("polymer_name")
        solv = entry.get("solvent")
        if not poly or poly == "N/A" or not solv or solv == "N/A":
            continue
        valid_flory.append(entry)
        
    records_flory = []
    for entry in valid_flory:
        records_flory.append({
            "solvent": entry.get("solvent"),
            "polymer": entry.get("polymer_name"),
            "c_value": entry.get("c_value"),
            "v_value": entry.get("v_value"),
            "source_paper": entry.get("source_paper"),
            "table_name": entry.get("table_name")
        })
    df_flory = pd.DataFrame(records_flory)
    if df_flory.empty:
        df_flory = pd.DataFrame(columns=["solvent", "polymer", "c_value", "v_value", "source_paper", "table_name"])
        
    # 2. Mark-Houwink flat database
    mh_entries = results.get("mark_houwink_entries", [])
    valid_mh = []
    for entry in mh_entries:
        if entry.get("failed_fields") and entry["failed_fields"] != "None":
            continue
        poly = entry.get("polymer_name")
        solv = entry.get("solvent")
        if not poly or poly == "N/A" or not solv or solv == "N/A":
            continue
        valid_mh.append(entry)
        
    records_mh = []
    for entry in valid_mh:
        records_mh.append({
            "solvent": entry.get("solvent"),
            "polymer": entry.get("polymer_name"),
            "K_value": entry.get("K_value"),
            "a_value": entry.get("a_value"),
            "source_paper": entry.get("source_paper"),
            "table_name": entry.get("table_name")
        })
    df_mh = pd.DataFrame(records_mh)
    if df_mh.empty:
        df_mh = pd.DataFrame(columns=["solvent", "polymer", "K_value", "a_value", "source_paper", "table_name"])
    
    # 5. Write outputs
    try:
        # Create output folders if they do not exist
        out_dir = os.path.dirname(json_out)
        if out_dir and not os.path.exists(out_dir):
            os.makedirs(out_dir, exist_ok=True)
            
        with open(json_out, 'w', encoding='utf-8') as jf:
            json.dump(results, jf, indent=2, ensure_ascii=False)
            
        # Save flat databases to file system
        df_flory.to_pickle(f"{output_prefix}_flory_database.pkl")
        df_flory.to_csv(f"{output_prefix}_flory_database.csv", index=False)
        df_mh.to_pickle(f"{output_prefix}_mh_database.pkl")
        df_mh.to_csv(f"{output_prefix}_mh_database.csv", index=False)
            
        create_combined_excel(xlsx_out, results)
        
    except Exception as e:
        console.print(f"[bold red]Error saving outputs: {e}[/bold red]")
        sys.exit(1)
        
    # 6. Report results and statistics
    console.print()
    console.print(f"[bold green]✓[/bold green] Combining process completed in [yellow]{elapsed:.2f}s[/yellow].")
    console.print(f"JSON data saved to: [cyan]{json_out}[/cyan]")
    console.print(f"Excel report saved to: [cyan]{xlsx_out}[/cyan]")
    console.print(f"Flory database saved to: [cyan]{output_prefix}_flory_database.pkl / .csv[/cyan]")
    console.print(f"Mark-Houwink database saved to: [cyan]{output_prefix}_mh_database.pkl / .csv[/cyan]")
    console.print()
    
    # Statistics Table
    stats = results.get("stats", {})
    t = Table(title="Homogenisation Statistics", header_style="bold magenta")
    t.add_column("Metric", style="cyan")
    t.add_column("Count", style="green")
    
    t.add_row("Total Chemical Fields Checked", str(stats.get("total_processed", 0)))
    t.add_row("Cache Hits (Fuzzy + Exact)", str(stats.get("cache_hits", 0)))
    t.add_row("AI Match Hits (Enum Selection)", str(stats.get("ai_match_hits", 0)))
    t.add_row("AI Generated Names (New)", str(stats.get("ai_generated", 0)))
    t.add_row("Failed / Non-Chemical Fields (N/A)", str(stats.get("failed", 0)))
    console.print(t)
    
    failures = results.get("failures", [])
    if failures:
        console.print()
        console.print(f"[bold yellow]Recorded Failures ({len(failures)}):[/bold yellow]")
        for f in failures[:15]:
            console.print(f" - [dim]{f['source_paper']}/{f['table']}[/dim]: Field [cyan]{f['field']}[/cyan] with value '[red]{f['value']}[/red]' -> [yellow]{f['reason']}[/yellow]")
        if len(failures) > 15:
            console.print(f" ... and {len(failures) - 15} more failures. See the failures sheet in Excel/JSON.")
            
    # Calculate token counts and pricing
    # (Since AI prompt increments these totals, we can report model total costs)
    total_in = ai_instance.total_prompt_tokens
    total_out = ai_instance.total_candidate_tokens
    if total_in > 0 or total_out > 0:
        cost_str = ""
        if selected_model in pricing_matrix:
            pricing = pricing_matrix[selected_model]
            cost = (total_in * pricing["input_per_m"] + total_out * pricing["output_per_m"]) / 1_000_000
            cost_str = f" ($ {cost:.6f})"
        console.print()
        console.print(f"AI Usage during run: [magenta]{selected_model}[/magenta] - [dim]{total_in} input tokens, {total_out} output tokens{cost_str}[/dim]")
