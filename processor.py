import os
import json
import csv
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import gc
import time
from rich.progress import Progress, SpinnerColumn, TextColumn, TimeElapsedColumn
from rich.console import Console
from rich.tree import Tree

from extractor import TableExtractor
from categorisation import categorise_table
from summariser import summarise_table_conditions, extract_paper_metadata


class PDFProcessor:
    def __init__(self, pdf_path: str, clean_dir: str, output_dir: str, logs_dir: str, model: str = "gemini"):
        self.pdf_path = pdf_path
        self.clean_dir = clean_dir
        self.output_dir = output_dir
        self.logs_dir = logs_dir
        self.model = model
        
        self.base_name = os.path.basename(pdf_path)
        self.base_no_ext = os.path.splitext(self.base_name)[0]
        
        # Output paths
        self.clean_path = os.path.join(self.clean_dir, f"clean_{self.base_name}")
        self.parsed_md_path = os.path.join(self.output_dir, "output.md")
        self.tables_dir = os.path.join(self.output_dir, "tables")
        self.log_file_path = os.path.join(self.logs_dir, f"log_{self.base_name}.log")
        
        # State
        self.extractor = None
        self.num_tables = 0
        self.cat_results = []
        self.sum_results = []
        self.metadata_res = None

    def extract(self, console=None):
        """Extracts content in-memory with a loading spinner and elapsed timer if console is provided."""
        
        start_time = time.time()
        if console:
            with Progress(
                SpinnerColumn(),
                TextColumn("[progress.description]{task.description}"),
                TimeElapsedColumn(),
                console=console,
                transient=True
            ) as progress:
                progress.add_task(f"[bold cyan]Extracting '{self.base_name}'[/bold cyan]...")
                self.extractor = TableExtractor(self.pdf_path)
        else:
            self.extractor = TableExtractor(self.pdf_path)
        self.num_tables = len(self.extractor.tables_markdown)

        # Save output files automatically
        self.save_cleaned_pdf()
        self.save_outputs()
        self.save_logs()

        elapsed_time = time.time() - start_time
        if console:
            self.print_extraction_tree(console, elapsed_time)

    def save_cleaned_pdf(self):
        """Saves the cleaned PDF byte content to the specified path."""
        if not self.extractor:
            raise RuntimeError("Must call extract() before saving cleaned PDF.")
        with open(self.clean_path, "wb") as clean_file:
            clean_file.write(self.extractor.clean_pdf_bytes)

    def save_outputs(self):
        """Saves the parsed markdown and extraction tables (text & csv)."""
        if not self.extractor:
            raise RuntimeError("Must call extract() before saving outputs.")
            
        # Save parsed markdown
        with open(self.parsed_md_path, "w", encoding="utf-8") as parsed_file:
            parsed_file.write(self.extractor.parsed_markdown)
            
        # Save text tables
        txt_dir = os.path.join(self.tables_dir, "txt")
        os.makedirs(txt_dir, exist_ok=True)
        for i, table_str in enumerate(self.extractor.tables_markdown):
            table_file_path = os.path.join(txt_dir, f"table{i + 1}.txt")
            with open(table_file_path, "w", encoding="utf-8") as table_file:
                table_file.write(table_str)
                
        # Save CSV tables
        csv_dir = os.path.join(self.tables_dir, "csv")
        os.makedirs(csv_dir, exist_ok=True)
        for i, csv_str in enumerate(self.extractor.tables_csv):
            csv_file_path = os.path.join(csv_dir, f"table{i + 1}.csv")
            with open(csv_file_path, "w", encoding="utf-8") as csv_file:
                csv_file.write(csv_str)

    def categorise_tables(self, console=None):
        """Categorises each saved table text file and saves JSON outputs."""
        if not self.extractor:
            raise RuntimeError("Must call extract() before categorising tables.")
            
        start_time = time.time()
        categorisation_dir = os.path.join(self.output_dir, "categorisation")
        
        def run_cat():
            os.makedirs(categorisation_dir, exist_ok=True)
            self.cat_results = []
            txt_dir = os.path.join(self.tables_dir, "txt")
            
            for i in range(self.num_tables):
                table_file_path = os.path.join(txt_dir, f"table{i + 1}.txt")
                table_name = os.path.basename(table_file_path)
                json_file_path = os.path.join(categorisation_dir, f"table{i + 1}.json")
                if os.path.exists(table_file_path):
                    res = categorise_table(table_file_path, model=self.model)
                    if res.success:
                        status = "Contains" if res.contains_diffusion else "Does NOT contain"
                        self.cat_results.append((table_name, True, f"{status} chemical diffusion coefficient data", res.usage_metadata))
                        
                        cat_data = {
                            "contains_scientific_data": res.contains_scientific_data,
                            "contains_diffusion_coeff": res.contains_diffusion_coeff,
                            "contains_polymer_diffusion_coeff": res.contains_polymer_diffusion_coeff,
                            "contains_diffusion": res.contains_diffusion
                        }
                        with open(json_file_path, 'w', encoding='utf-8') as jf:
                            json.dump(cat_data, jf, indent=2)
                    else:
                        self.cat_results.append((table_name, False, f"Failed: {res.error}", None))
                        
        if console:
            with Progress(
                SpinnerColumn(),
                TextColumn("[progress.description]{task.description}"),
                TimeElapsedColumn(),
                console=console,
                transient=True
            ) as progress:
                progress.add_task(f"[bold cyan]Categorising '{self.base_name}'[/bold cyan]...")
                run_cat()
            elapsed_time = time.time() - start_time
            self.print_categorisation_tree(console, elapsed_time)
        else:
            run_cat()

    def summarise_tables(self, console=None):
        """Summarises each saved table text file and saves JSON outputs to descriptions_dir."""
        
        if not self.extractor:
            raise RuntimeError("Must call extract() before summarising tables.")
            
        start_time = time.time()
        descriptions_dir = os.path.join(self.output_dir, "descriptions")
        
        def run_summarise():
            os.makedirs(descriptions_dir, exist_ok=True)
            self.sum_results = []
            
            # Extract paper-level metadata
            metadata_res = extract_paper_metadata(self.extractor.parsed_markdown, model=self.model)
            self.metadata_res = metadata_res
            
            metadata_dict = {}
            metadata_error = None
            if metadata_res.success:
                metadata_dict = metadata_res.data.model_dump()
            else:
                metadata_error = metadata_res.error
                
            txt_dir = os.path.join(self.tables_dir, "txt")
            for i in range(self.num_tables):
                table_file_path = os.path.join(txt_dir, f"table{i + 1}.txt")
                table_name = os.path.basename(table_file_path)
                json_file_path = os.path.join(descriptions_dir, f"table{i + 1}.json")
                
                if os.path.exists(table_file_path):
                    with open(table_file_path, 'r', encoding='utf-8') as f:
                        table_text = f.read()
                        
                    res = summarise_table_conditions(table_text, model=self.model)
                    
                    if res.success:
                        # Merge paper metadata and experimental conditions
                        combined_data = {
                            **metadata_dict,
                            **res.data.model_dump()
                        }

                        with open(json_file_path, 'w', encoding='utf-8') as jf:
                            json.dump(combined_data, jf, indent=2)
                        
                        status_msg = "Successfully summarised"
                        if metadata_error:
                            status_msg += f" (metadata extraction failed: {metadata_error})"
                        self.sum_results.append((table_name, True, status_msg, res.usage_metadata))
                    else:
                        self.sum_results.append((table_name, False, f"Failed: {res.error}", None))
                        
        if console:
            with Progress(
                SpinnerColumn(),
                TextColumn("[progress.description]{task.description}"),
                TimeElapsedColumn(),
                console=console,
                transient=True
            ) as progress:
                progress.add_task(f"[bold cyan]Summarising '{self.base_name}'[/bold cyan]...")
                run_summarise()
            elapsed_time = time.time() - start_time
            self.print_summarisation_tree(console, elapsed_time)
        else:
            run_summarise()

    def create_excel(self) -> None:
        """Creates a beautifully formatted Excel document containing both JSON metadata/conditions and CSV table data."""
        wb = openpyxl.Workbook()
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])

        font_family = "Segoe UI"
        
        title_font = Font(name=font_family, size=16, bold=True, color="1F497D")
        title_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
        title_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        section_font = Font(name=font_family, size=11, bold=True, color="1F497D")
        section_fill = PatternFill(start_color="B8CCE4", end_color="B8CCE4", fill_type="solid")
        
        label_font = Font(name=font_family, size=10, bold=True, color="333333")
        val_font = Font(name=font_family, size=10, color="000000")
        desc_val_font = Font(name=font_family, size=10, italic=True, color="333333")
        
        table_header_font = Font(name=font_family, size=10, bold=True, color="FFFFFF")
        table_header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        table_header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        thin_border_side = Side(border_style="thin", color="D9D9D9")
        data_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
        thick_bottom = Border(bottom=Side(border_style="medium", color="1F497D"))
        
        left_align = Alignment(horizontal="left", vertical="center")
        right_align = Alignment(horizontal="right", vertical="center")
        wrap_left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)

        for i in range(self.num_tables):
            sheet_name = f"Table {i + 1}"
            ws = wb.create_sheet(title=sheet_name)
            ws.views.sheetView[0].showGridLines = True
            
            json_path = os.path.join(self.output_dir, "descriptions", f"table{i + 1}.json")
            csv_path = os.path.join(self.tables_dir, "csv", f"table{i + 1}.csv")
            
            data = {}
            if os.path.exists(json_path):
                try:
                    with open(json_path, 'r', encoding='utf-8') as jf:
                        data = json.load(jf)
                except Exception as e:
                    print(f"Error loading JSON {json_path}: {e}")
            
            title_text = data.get("title", self.base_no_ext)
            ws.merge_cells("A1:G2")
            title_cell = ws["A1"]
            title_cell.value = title_text
            title_cell.font = title_font
            title_cell.fill = title_fill
            title_cell.alignment = title_align
            
            for row in range(1, 3):
                for col in range(1, 8):
                    cell = ws.cell(row=row, column=col)
                    cell.fill = title_fill
                    
            ws.cell(row=4, column=1, value="PAPER METADATA & EXPERIMENTAL CONDITIONS").font = section_font
            ws.merge_cells("A4:G4")
            for col in range(1, 8):
                cell = ws.cell(row=4, column=col)
                cell.fill = section_fill
                cell.border = thick_bottom

            def write_metadata_row(r, label, value, is_description=False):
                ws.cell(row=r, column=1, value=label).font = label_font
                ws.cell(row=r, column=1).alignment = left_align
                ws.cell(row=r, column=2, value=value).font = desc_val_font if is_description else val_font
                ws.cell(row=r, column=2).alignment = wrap_left_align
                ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=7)
                for col in range(1, 8):
                    ws.cell(row=r, column=col).border = Border(bottom=thin_border_side)
            
            authors_val = ", ".join(data.get("authors", [])) if isinstance(data.get("authors"), list) else data.get("authors", "")
            write_metadata_row(5, "Authors", authors_val)
            write_metadata_row(6, "DOI", data.get("doi", ""))
            write_metadata_row(7, "Temperature", data.get("temperature", ""))
            write_metadata_row(8, "Pressure", data.get("pressure", ""))
            chemicals_val = ", ".join(data.get("chemicals", [])) if isinstance(data.get("chemicals"), list) else data.get("chemicals", "")
            write_metadata_row(9, "Chemicals", chemicals_val)
            
            desc_val = data.get("description", "")
            write_metadata_row(10, "Description", desc_val, is_description=True)
            ws.row_dimensions[10].height = 45
            
            other_stats = data.get("other_statistics", [])
            stats_str = ""
            if isinstance(other_stats, list):
                stats_str = ", ".join([f"{stat.get('name')}: {stat.get('value')}" for stat in other_stats if isinstance(stat, dict)])
            elif isinstance(other_stats, dict):
                stats_str = ", ".join([f"{k}: {v}" for k, v in other_stats.items()])
            write_metadata_row(11, "Other Stats", stats_str)

            cat_path = os.path.join(self.output_dir, "categorisation", f"table{i + 1}.json")
            cat_data = {}
            if os.path.exists(cat_path):
                try:
                    with open(cat_path, 'r', encoding='utf-8') as cf:
                        cat_data = json.load(cf)
                except Exception as e:
                    print(f"Error loading categorisation JSON {cat_path}: {e}")

            curr_row = 13
            if cat_data:
                ws.cell(row=curr_row, column=1, value="TABLE CATEGORISATION").font = section_font
                ws.merge_cells(start_row=curr_row, start_column=1, end_row=curr_row, end_column=7)
                for col in range(1, 8):
                    cell = ws.cell(row=curr_row, column=col)
                    cell.fill = section_fill
                    cell.border = thick_bottom
                
                curr_row += 1
                def format_bool(val):
                    if val is True:
                        return "Yes"
                    if val is False:
                        return "No"
                    return ""
                write_metadata_row(curr_row, "Contains Scientific Data", format_bool(cat_data.get("contains_scientific_data")))
                curr_row += 1
                write_metadata_row(curr_row, "Contains Diffusion Coefficients", format_bool(cat_data.get("contains_diffusion_coeff")))
                curr_row += 1
                write_metadata_row(curr_row, "Contains Polymer Diffusion Coefficients", format_bool(cat_data.get("contains_polymer_diffusion_coeff")))
                
                curr_row += 2 # gap + dynamic next section start

            start_row = curr_row
            ws.cell(row=start_row, column=1, value="EXTRACTED TABLE DATA").font = section_font
            ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=7)
            for col in range(1, 8):
                cell = ws.cell(row=start_row, column=col)
                cell.fill = section_fill
                cell.border = thick_bottom
                
            csv_start_row = start_row + 1
            
            if os.path.exists(csv_path):
                try:
                    with open(csv_path, 'r', encoding='utf-8', newline='') as cf:
                        reader = csv.reader(cf)
                        csv_rows = list(reader)
                    
                    active_row_idx = 0
                    for csv_row in csv_rows:
                        # Skip empty rows (blank line or list containing only whitespace strings)
                        if not csv_row or all(val.strip() == '' for val in csv_row):
                            continue
                            
                        curr_row_idx = csv_start_row + active_row_idx
                        is_header = (active_row_idx == 0)
                        
                        for c_idx, val in enumerate(csv_row):
                            cell = ws.cell(row=curr_row_idx, column=c_idx + 1)
                            
                            parsed_val = val
                            try:
                                if "." in val:
                                    parsed_val = float(val)
                                else:
                                    parsed_val = int(val)
                            except ValueError:
                                pass
                                
                            cell.value = parsed_val
                            cell.border = data_border
                            
                            if is_header:
                                cell.font = table_header_font
                                cell.fill = table_header_fill
                                cell.alignment = table_header_align
                            else:
                                cell.font = val_font
                                if isinstance(parsed_val, (int, float)):
                                    cell.alignment = right_align
                                else:
                                    cell.alignment = left_align
                            active_row_idx += 1
                except Exception as e:
                    print(f"Error parsing CSV {csv_path}: {e}")
                    ws.cell(row=csv_start_row, column=1, value=f"Error loading CSV data: {e}").font = val_font
            else:
                ws.cell(row=csv_start_row, column=1, value="CSV data file not found.").font = val_font

            excluded_rows = {1, 2, 4, 10, start_row}
            if cat_data:
                excluded_rows.add(13)

            for col in ws.columns:
                max_len = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    if cell.row in excluded_rows:
                        continue
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
                ws.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 40)

        output_filename = f"{self.base_no_ext}_summary.xlsx"
        dest_path = os.path.join(self.output_dir, output_filename)
        try:
            wb.save(dest_path)
        except Exception as e:
            print(f"Error saving Excel document {dest_path}: {e}")

    def save_logs(self):
        """Saves the extraction log file."""
        if not self.extractor:
            raise RuntimeError("Must call extract() before saving logs.")
        with open(self.log_file_path, "w", encoding="utf-8") as log_file:
            log_file.write(self.extractor.logs)

    def print_extraction_tree(self, console: Console, elapsed_time: float):
        """Builds and prints the hierarchical tree status for the extraction phase."""
        tree = Tree(f"[bold cyan]📄 {self.base_name}[/bold cyan] [dim](Extraction completed in {elapsed_time:.2f}s)[/dim]")
        tree.add(f"[green]✓[/green] Extracted text & tables in-memory")
        tree.add(f"[green]✓[/green] Saved cleaned PDF to [yellow]{os.path.relpath(self.clean_path)}[/yellow]")
        tree.add(f"[green]✓[/green] Saved parsed markdown to [yellow]{os.path.relpath(self.parsed_md_path)}[/yellow]")
        tree.add(f"[green]✓[/green] Saved {self.num_tables} tables (txt & csv) to [yellow]{os.path.relpath(self.tables_dir)}[/yellow]")
        tree.add(f"[green]✓[/green] Saved execution logs to [yellow]{os.path.relpath(self.log_file_path)}[/yellow]")
        console.print(tree)
        console.print()

    def print_categorisation_tree(self, console: Console, elapsed_time: float):
        """Builds and prints the hierarchical tree status for the table categorisation phase."""
        total_in = 0
        total_out = 0
        has_tokens = False

        if self.cat_results:
            for item in self.cat_results:
                if len(item) == 4 and item[3]:
                    total_in += item[3].get("prompt_token_count", 0)
                    total_out += item[3].get("candidates_token_count", 0)
                    has_tokens = True

        tokens_title = ""
        if has_tokens:
            tokens_title = f" [dim](Total tokens: {total_in} in, {total_out} out)[/dim]"

        tree = Tree(f"[bold cyan]📄 {self.base_name}[/bold cyan] [dim](Categorisation completed in {elapsed_time:.2f}s)[/dim]{tokens_title}")
        cat_node = tree.add(f"[green]✓[/green] Categorised extracted tables using [magenta]{self.model}[/magenta]")
        
        if not self.cat_results:
            cat_node.add("[dim]No tables found to categorise[/dim]")
        else:
            for item in self.cat_results:
                table_name = item[0]
                success = item[1]
                status = item[2]
                usage_metadata = item[3] if len(item) == 4 else None

                tokens_str = ""
                if usage_metadata:
                    in_t = usage_metadata.get("prompt_token_count", 0)
                    out_t = usage_metadata.get("candidates_token_count", 0)
                    tokens_str = f" [dim](tokens: {in_t} in, {out_t} out)[/dim]"

                if success:
                    if "Does NOT contain" in status:
                        status_styled = f"[blue]{status}[/blue]"
                    else:
                        status_styled = f"[bold green]{status}[/bold green]"
                    cat_node.add(f"{table_name}: {status_styled}{tokens_str}")
                else:
                    cat_node.add(f"{table_name}: [red]{status}[/red]")
        
        console.print(tree)
        console.print()

    def print_summarisation_tree(self, console: Console, elapsed_time: float):
        """Builds and prints the hierarchical tree status for the table summarisation phase."""
        total_in = 0
        total_out = 0
        has_tokens = False

        if self.metadata_res and self.metadata_res.success and self.metadata_res.usage_metadata:
            total_in += self.metadata_res.usage_metadata.get("prompt_token_count", 0)
            total_out += self.metadata_res.usage_metadata.get("candidates_token_count", 0)
            has_tokens = True
        if self.sum_results:
            for item in self.sum_results:
                if len(item) == 4 and item[3]:
                    total_in += item[3].get("prompt_token_count", 0)
                    total_out += item[3].get("candidates_token_count", 0)
                    has_tokens = True

        tokens_title = ""
        if has_tokens:
            tokens_title = f" [dim](Total tokens: {total_in} in, {total_out} out)[/dim]"

        tree = Tree(f"[bold cyan]📄 {self.base_name}[/bold cyan] [dim](Summary completed in {elapsed_time:.2f}s)[/dim]{tokens_title}")
        sum_node = tree.add(f"[green]✓[/green] Summarised experimental conditions using [magenta]{self.model}[/magenta]")
        
        if self.metadata_res:
            tokens_str = ""
            if self.metadata_res.usage_metadata:
                in_t = self.metadata_res.usage_metadata.get("prompt_token_count", 0)
                out_t = self.metadata_res.usage_metadata.get("candidates_token_count", 0)
                tokens_str = f" [dim](tokens: {in_t} in, {out_t} out)[/dim]"
            
            if self.metadata_res.success:
                sum_node.add(f"Paper Metadata: [bold green]Successfully extracted[/bold green]{tokens_str}")
            else:
                sum_node.add(f"Paper Metadata: [red]Failed: {self.metadata_res.error}[/red]")

        if not self.sum_results:
            sum_node.add("[dim]No tables found to summarise[/dim]")
        else:
            for item in self.sum_results:
                table_name = item[0]
                success = item[1]
                status = item[2]
                usage_metadata = item[3] if len(item) == 4 else None

                tokens_str = ""
                if usage_metadata:
                    in_t = usage_metadata.get("prompt_token_count", 0)
                    out_t = usage_metadata.get("candidates_token_count", 0)
                    tokens_str = f" [dim](tokens: {in_t} in, {out_t} out)[/dim]"

                if success:
                    sum_node.add(f"{table_name}: [bold green]{status}[/bold green]{tokens_str}")
                else:
                    sum_node.add(f"{table_name}: [red]{status}[/red]")
        
        console.print(tree)
        console.print()

    def cleanup(self):
        """Deload extractor instance and force garbage collection to release memory."""
        self.extractor = None
        gc.collect()
