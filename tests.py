import sys
## Allows unicode in std out
if hasattr(sys.stdout, 'reconfigure'):
    sys.stdout.reconfigure(encoding='utf-8')
if hasattr(sys.stderr, 'reconfigure'):
    sys.stderr.reconfigure(encoding='utf-8')

import glob
import os
import json
import csv
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import time
import gc

from rich.console import Console
from rich.progress import Progress, SpinnerColumn, TextColumn, TimeElapsedColumn
from rich.tree import Tree
from rich.rule import Rule
from rich.table import Table

from extractor import TableExtractor
from categorisation import categorise_table
from summariser import summarise_table_conditions, extract_paper_metadata

TEST_DIR = "./tests"
MATERIAL_DIR = "./tests/material"


def setup_run_layout(test_dir: str) -> tuple[str, str, str]:
    """Creates the timestamped directories for the run and returns their paths."""
    run_id = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    run_dir = os.path.join(test_dir, "runs", run_id)
    
    clean_dir = os.path.join(run_dir, "clean")
    output_dir = os.path.join(run_dir, "output")
    logs_dir = os.path.join(run_dir, "logs")
    
    os.makedirs(clean_dir, exist_ok=True)
    os.makedirs(output_dir, exist_ok=True)
    os.makedirs(logs_dir, exist_ok=True)
    
    return clean_dir, output_dir, logs_dir


## ------ Save files to folders -------

def save_cleaned_pdf(clean_path: str, clean_pdf_bytes: bytes):
    """Saves the cleaned PDF byte content to the specified path."""
    with open(clean_path, "wb") as clean_file:
        clean_file.write(clean_pdf_bytes)


def save_parsed_markdown(parsed_md_path: str, parsed_markdown: str):
    """Saves the entire parsed markdown to the specified path."""
    with open(parsed_md_path, "w", encoding="utf-8") as parsed_file:
        parsed_file.write(parsed_markdown)


def save_tables_text(tables_dir: str, tables_list: list[str]):
    """Saves each extracted table string to table1.txt, table2.txt, etc. in a 'txt' subfolder within tables_dir."""
    txt_dir = os.path.join(tables_dir, "txt")
    os.makedirs(txt_dir, exist_ok=True)
    for i, table_str in enumerate(tables_list):
        table_file_path = os.path.join(txt_dir, f"table{i + 1}.txt")
        with open(table_file_path, "w", encoding="utf-8") as table_file:
            table_file.write(table_str)


def save_tables_csv(tables_dir: str, csv_tables: list[str]):
    """Saves each extracted table CSV string to table1.csv, table2.csv, etc. in a 'csv' subfolder within tables_dir."""
    csv_dir = os.path.join(tables_dir, "csv")
    os.makedirs(csv_dir, exist_ok=True)
    for i, csv_str in enumerate(csv_tables):
        csv_file_path = os.path.join(csv_dir, f"table{i + 1}.csv")
        with open(csv_file_path, "w", encoding="utf-8") as csv_file:
            csv_file.write(csv_str)


def save_logs(logs_dir: str, base_name: str, logs_content: str):
    """Saves the extraction log file."""
    log_file_path = os.path.join(logs_dir, f"log_{base_name}.log")
    with open(log_file_path, "w", encoding="utf-8") as log_file:
        log_file.write(logs_content)


def categorise_extracted_tables(tables_dir: str, num_tables: int, model: str) -> list[tuple[str, bool, str]]:
    """Categorises each saved table text file and returns a list of results."""
    results = []
    for i in range(num_tables):
        table_file_path = os.path.join(tables_dir, "txt", f"table{i + 1}.txt")
        table_name = os.path.basename(table_file_path)
        if os.path.exists(table_file_path):
            res = categorise_table(table_file_path, model=model)
            if res.success:
                status = "Contains" if res.contains_diffusion else "Does NOT contain"
                results.append((table_name, True, f"{status} chemical diffusion coefficient data"))
            else:
                results.append((table_name, False, f"Failed: {res.error}"))
    return results


def summarise_extracted_tables(tables_dir: str, descriptions_dir: str, num_tables: int, parsed_markdown: str, model: str) -> list[tuple[str, bool, str]]:
    """Summarises each saved table text file and saves JSON outputs to descriptions_dir."""
    os.makedirs(descriptions_dir, exist_ok=True)
    results = []
    
    # 1. Extract paper-level metadata
    metadata_res = extract_paper_metadata(parsed_markdown, model=model)
    metadata_dict = {}
    metadata_error = None
    if metadata_res.success:
        metadata_dict = metadata_res.data.model_dump()
    else:
        metadata_error = metadata_res.error
        
    for i in range(num_tables):
        table_file_path = os.path.join(tables_dir, "txt", f"table{i + 1}.txt")
        table_name = os.path.basename(table_file_path)
        json_file_path = os.path.join(descriptions_dir, f"table{i + 1}.json")
        
        if os.path.exists(table_file_path):
            with open(table_file_path, 'r', encoding='utf-8') as f:
                table_text = f.read()
                
            res = summarise_table_conditions(table_text, model=model)
            if res.success:
                # Merge paper metadata and experimental conditions
                combined_data = {
                    **metadata_dict,
                    **res.data.model_dump()
                }
                with open(json_file_path, 'w', encoding='utf-8') as jf:
                    json.dump(combined_data, jf, indent=2)
                
                status_msg = "Successfully summarised"
                if metadata_error:
                    status_msg += f" (metadata failed: {metadata_error})"
                results.append((table_name, True, status_msg))
            else:
                results.append((table_name, False, f"Failed: {res.error}"))
    return results


def create_excel(test_output_dir: str, base_name: str, num_tables: int) -> None:
    """Creates a beautifully formatted Excel document containing both JSON metadata/conditions and CSV table data."""
    wb = openpyxl.Workbook()
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    font_family = "Segoe UI"
    
    title_font = Font(name=font_family, size=16, bold=True, color="1F497D")
    title_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
    title_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    section_font = Font(name=font_family, size=11, bold=True, color="1F497D")
    section_fill = PatternFill(start_color="B8CCE4", end_color="B8CCE4", fill_type="solid")
    
    label_font = Font(name=font_family, size=10, bold=True, color="333333")
    val_font = Font(name=font_family, size=10, color="000000")
    desc_val_font = Font(name=font_family, size=10, italic=True, color="333333")
    
    table_header_font = Font(name=font_family, size=10, bold=True, color="FFFFFF")
    table_header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    table_header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    thin_border_side = Side(border_style="thin", color="D9D9D9")
    data_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    thick_bottom = Border(bottom=Side(border_style="medium", color="1F497D"))
    
    left_align = Alignment(horizontal="left", vertical="center")
    right_align = Alignment(horizontal="right", vertical="center")
    wrap_left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)

    for i in range(num_tables):
        sheet_name = f"Table {i + 1}"
        ws = wb.create_sheet(title=sheet_name)
        ws.views.sheetView[0].showGridLines = True
        
        json_path = os.path.join(test_output_dir, "descriptions", f"table{i + 1}.json")
        csv_path = os.path.join(test_output_dir, "tables", "csv", f"table{i + 1}.csv")
        
        data = {}
        if os.path.exists(json_path):
            try:
                with open(json_path, 'r', encoding='utf-8') as jf:
                    data = json.load(jf)
            except Exception as e:
                print(f"Error loading JSON {json_path}: {e}")
        
        title_text = data.get("title", base_name)
        ws.merge_cells("A1:G2")
        title_cell = ws["A1"]
        title_cell.value = title_text
        title_cell.font = title_font
        title_cell.fill = title_fill
        title_cell.alignment = title_align
        
        for row in range(1, 3):
            for col in range(1, 8):
                cell = ws.cell(row=row, column=col)
                cell.fill = title_fill
                
        ws.cell(row=4, column=1, value="PAPER METADATA & EXPERIMENTAL CONDITIONS").font = section_font
        ws.merge_cells("A4:G4")
        for col in range(1, 8):
            cell = ws.cell(row=4, column=col)
            cell.fill = section_fill
            cell.border = thick_bottom

        def write_metadata_row(r, label, value, is_description=False):
            ws.cell(row=r, column=1, value=label).font = label_font
            ws.cell(row=r, column=1).alignment = left_align
            ws.cell(row=r, column=2, value=value).font = desc_val_font if is_description else val_font
            ws.cell(row=r, column=2).alignment = wrap_left_align
            ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=7)
            for col in range(1, 8):
                ws.cell(row=r, column=col).border = Border(bottom=thin_border_side)
        
        authors_val = ", ".join(data.get("authors", [])) if isinstance(data.get("authors"), list) else data.get("authors", "")
        write_metadata_row(5, "Authors", authors_val)
        write_metadata_row(6, "DOI", data.get("doi", ""))
        write_metadata_row(7, "Temperature", data.get("temperature", ""))
        write_metadata_row(8, "Pressure", data.get("pressure", ""))
        chemicals_val = ", ".join(data.get("chemicals", [])) if isinstance(data.get("chemicals"), list) else data.get("chemicals", "")
        write_metadata_row(9, "Chemicals", chemicals_val)
        
        desc_val = data.get("description", "")
        write_metadata_row(10, "Description", desc_val, is_description=True)
        ws.row_dimensions[10].height = 45
        
        other_stats = data.get("other_statistics", [])
        stats_str = ""
        if isinstance(other_stats, list):
            stats_str = ", ".join([f"{stat.get('name')}: {stat.get('value')}" for stat in other_stats if isinstance(stat, dict)])
        elif isinstance(other_stats, dict):
            stats_str = ", ".join([f"{k}: {v}" for k, v in other_stats.items()])
        write_metadata_row(11, "Other Stats", stats_str)

        start_row = 13
        ws.cell(row=start_row, column=1, value="EXTRACTED TABLE DATA").font = section_font
        ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=7)
        for col in range(1, 8):
            cell = ws.cell(row=start_row, column=col)
            cell.fill = section_fill
            cell.border = thick_bottom
            
        csv_start_row = start_row + 1
        
        if os.path.exists(csv_path):
            try:
                with open(csv_path, 'r', encoding='utf-8', newline='') as cf:
                    reader = csv.reader(cf)
                    csv_rows = list(reader)
                
                active_row_idx = 0
                for csv_row in csv_rows:
                    # Skip empty rows (blank line or list containing only whitespace strings)
                    if not csv_row or all(val.strip() == '' for val in csv_row):
                        continue
                        
                    curr_row = csv_start_row + active_row_idx
                    is_header = (active_row_idx == 0)
                    
                    for c_idx, val in enumerate(csv_row):
                        cell = ws.cell(row=curr_row, column=c_idx + 1)
                        
                        parsed_val = val
                        try:
                            if "." in val:
                                parsed_val = float(val)
                            else:
                                parsed_val = int(val)
                        except ValueError:
                            pass
                            
                        cell.value = parsed_val
                        cell.border = data_border
                        
                        if is_header:
                            cell.font = table_header_font
                            cell.fill = table_header_fill
                            cell.alignment = table_header_align
                        else:
                            cell.font = val_font
                            if isinstance(parsed_val, (int, float)):
                                cell.alignment = right_align
                            else:
                                cell.alignment = left_align
                    active_row_idx += 1
            except Exception as e:
                print(f"Error parsing CSV {csv_path}: {e}")
                ws.cell(row=csv_start_row, column=1, value=f"Error loading CSV data: {e}").font = val_font
        else:
            ws.cell(row=csv_start_row, column=1, value="CSV data file not found.").font = val_font

        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.row in [1, 2, 4, 10, 13]:
                    continue
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 40)

    output_filename = f"{os.path.basename(test_output_dir)}_summary.xlsx"
    dest_path = os.path.join(test_output_dir, output_filename)
    try:
        wb.save(dest_path)
    except Exception as e:
        print(f"Error saving Excel document {dest_path}: {e}")


## ------- Console ouputs -------

def print_pdf_run_tree(
    console: Console,
    base_name: str,
    elapsed_time: float,
    clean_path: str,
    parsed_md_path: str,
    tables_dir: str,
    num_tables: int,
    log_file_path: str,
    categorise_tables: bool = False,
    summarise_tables: bool = False,
    model: str = "",
    cat_results: list = None,
    sum_results: list = None
):
    """Builds and prints the hierarchical tree status for a specific test run."""
    tree = Tree(f"[bold cyan]📄 {base_name}[/bold cyan] [dim](Completed in {elapsed_time:.2f}s)[/dim]")
    tree.add(f"[green]✓[/green] Extracted text & tables in-memory")
    tree.add(f"[green]✓[/green] Saved cleaned PDF to [yellow]{os.path.relpath(clean_path)}[/yellow]")
    tree.add(f"[green]✓[/green] Saved parsed markdown to [yellow]{os.path.relpath(parsed_md_path)}[/yellow]")
    tree.add(f"[green]✓[/green] Saved {num_tables} tables (txt & csv) to [yellow]{os.path.relpath(tables_dir)}[/yellow]")
    
    if categorise_tables:
        cat_node = tree.add(f"[green]✓[/green] Categorised extracted tables using [magenta]{model}[/magenta]")
        if not cat_results:
            cat_node.add("[dim]No tables found to categorise[/dim]")
        else:
            for table_name, success, status in cat_results:
                if success:
                    if "Does NOT contain" in status:
                        status_styled = f"[blue]{status}[/blue]"
                    else:
                        status_styled = f"[bold green]{status}[/bold green]"
                    cat_node.add(f"{table_name}: {status_styled}")
                else:
                    cat_node.add(f"{table_name}: [red]{status}[/red]")

    if summarise_tables:
        sum_node = tree.add(f"[green]✓[/green] Summarised experimental conditions using [magenta]{model}[/magenta]")
        if not sum_results:
            sum_node.add("[dim]No tables found to summarise[/dim]")
        else:
            for table_name, success, status in sum_results:
                if success:
                    sum_node.add(f"{table_name}: [bold green]{status}[/bold green]")
                else:
                    sum_node.add(f"{table_name}: [red]{status}[/red]")
                    
    tree.add(f"[green]✓[/green] Saved execution logs to [yellow]{os.path.relpath(log_file_path)}[/yellow]")
    
    console.print(tree)
    console.print()


def print_summary_table(console: Console, summary_data: list[dict]):
    """Prints the final summary table of all PDF runs."""
    console.print(Rule(title="[bold green]TEST SUITE SUMMARY[/bold green]", style="green"))
    console.print()
    table = Table(show_header=True, header_style="bold magenta")
    table.add_column("PDF File", style="cyan")
    table.add_column("Tables Extracted", justify="right", style="green")
    table.add_column("Execution Time", justify="right", style="yellow")
    
    total_time = 0.0
    total_tables = 0
    for row in summary_data:
        table.add_row(
            row["file"],
            str(row["tables"]),
            f"{row['time']:.2f}s"
        )
        total_time += row["time"]
        total_tables += row["tables"]
        
    table.add_section()
    table.add_row(
        "[bold]Total[/bold]",
        f"[bold]{total_tables}[/bold]",
        f"[bold]{total_time:.2f}s[/bold]"
    )
    console.print(table)
    console.print()


## -------- Main test script --------

def run_tests(categorise_tables: bool = True, summarise_tables: bool = True, model: str = "gemini"):
    # Initialize rich console targeting original stdout to bypass redirections
    console = Console(file=sys.__stdout__)
    
    # Find all test PDF files
    pdf_files = glob.glob(os.path.join(MATERIAL_DIR, "*.pdf"))
    
    console.print()
    console.print(Rule(title=f"[bold magenta]TESTING {len(pdf_files)} FILES[/bold magenta]", style="magenta"))
    console.print()
    
    # 1. Setup timestamped run directories
    clean_dir, output_dir, logs_dir = setup_run_layout(TEST_DIR)
    
    summary_data = []
    
    for pdf_path in pdf_files:
        base_name = os.path.basename(pdf_path)
        base_no_ext = os.path.splitext(base_name)[0]
        
        # Configure output paths within the timestamped run directory
        test_output_dir = os.path.join(output_dir, base_no_ext)
        os.makedirs(test_output_dir, exist_ok=True)
        clean_path = os.path.join(clean_dir, f"clean_{base_name}")
        
        timer = time.time()
        
        # 2. Extract content in-memory with a loading spinner and elapsed timer
        with Progress(
            SpinnerColumn(),
            TextColumn("[progress.description]{task.description}"),
            TimeElapsedColumn(),
            console=console,
            transient=True
        ) as progress:
            progress.add_task(f"[bold cyan]Extracting '{base_name}'[/bold cyan]...")
            extractor = TableExtractor(pdf_path)
            
        # 3. Save Cleaned PDF
        save_cleaned_pdf(clean_path, extractor.clean_pdf_bytes)
        
        # 4. Save Extraction Outputs (Markdown, Tables, CSVs)
        parsed_md_path = os.path.join(test_output_dir, "output.md")
        save_parsed_markdown(parsed_md_path, extractor.parsed_markdown)
        
        tables_dir = os.path.join(test_output_dir, "tables")
        save_tables_text(tables_dir, extractor.tables_markdown)
        save_tables_csv(tables_dir, extractor.tables_csv)
        
        # 5. Run Categorisation if enabled
        cat_results = []
        if categorise_tables:
            cat_results = categorise_extracted_tables(tables_dir, len(extractor.tables_markdown), model)
            
        # 5b. Run Summarisation if enabled
        sum_results = []
        if summarise_tables:
            descriptions_dir = os.path.join(test_output_dir, "descriptions")
            sum_results = summarise_extracted_tables(tables_dir, descriptions_dir, len(extractor.tables_markdown), extractor.parsed_markdown, model)
            
            # Generate Excel sheet
            create_excel(test_output_dir, base_no_ext, len(extractor.tables_markdown))
            
        # 6. Save Logs
        save_logs(logs_dir, base_name, extractor.logs)
        log_file_path = os.path.join(logs_dir, f"log_{base_name}.log")
        
        elapsed_time = time.time() - timer
        num_tables = len(extractor.tables_markdown)
        
        # Build and print tree list for this PDF file
        print_pdf_run_tree(
            console=console,
            base_name=base_name,
            elapsed_time=elapsed_time,
            clean_path=clean_path,
            parsed_md_path=parsed_md_path,
            tables_dir=tables_dir,
            num_tables=num_tables,
            log_file_path=log_file_path,
            categorise_tables=categorise_tables,
            summarise_tables=summarise_tables,
            model=model,
            cat_results=cat_results,
            sum_results=sum_results
        )
        
        # Collect data for final summary table
        summary_data.append({
            "file": base_name,
            "tables": num_tables,
            "time": elapsed_time
        })

        # 7. Deload extractor instance to release memory
        extractor = None
        gc.collect()
        
    # Print the final summary table
    print_summary_table(console, summary_data)


if __name__ == "__main__":    
    run_tests(model="gemini", categorise_tables=False, summarise_tables=True)